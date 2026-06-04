from __future__ import annotations

import base64
import csv
import difflib
import json
import logging
import unicodedata
from html import escape
from io import BytesIO, StringIO
from http.cookies import SimpleCookie
from datetime import date, datetime, timedelta
from decimal import Decimal
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from typing import Any
from urllib.parse import parse_qs, urlparse

import psycopg2
import psycopg2.extras

from web.backend.config import DATABASE_URL, EMPRESAS
from web.backend.auth.services.auth_service import AuthError, AuthService, PermissionDenied
from web.backend.modules import discover_routes
from web.backend.routing import RequestContext

try:
    import openpyxl
except Exception:
    openpyxl = None

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
except Exception:
    A4 = landscape = colors = ParagraphStyle = getSampleStyleSheet = None
    Paragraph = SimpleDocTemplate = Spacer = Table = TableStyle = None


HOST = "0.0.0.0"
PORT = 8008
SUCURSALES = {
    "aregua": {"local": "AREGUA", "nombre": "Aregua"},
    "luque": {"local": "LUQUE", "nombre": "Luque"},
    "itaugua": {"local": "ITAUGUA", "nombre": "Itaugua"},
}


def _json_default(value: Any):
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return str(value)


def _is_client_disconnect(exc: OSError) -> bool:
    return isinstance(exc, (BrokenPipeError, ConnectionAbortedError, ConnectionResetError)) or getattr(exc, "winerror", None) in {
        10053,
        10054,
    }


def _parse_date(value: str | None):
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%d").date()


def _parse_flexible_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _normalize_import_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return "".join(ch if ch.isalnum() else " " for ch in text).strip()


def _normalize_vehicle_import_ref(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return "".join(ch for ch in text if ch.isalnum())


def _normalize_invoice_number(value: Any) -> str:
    return "".join(str(value or "").strip().lower().split())


def _normalize_combustible_product(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("-", " ").replace("_", " ")
    return " ".join(text.split())


def _is_combustible_import_product(value: Any) -> bool:
    normalized = _normalize_combustible_product(value)
    if not normalized:
        return True
    excluded_tokens = ("atf", "aceite", "lubricante", "grasa", "filtro", "aditivo")
    if any(token in normalized for token in excluded_tokens):
        return False
    accepted_tokens = ("diesel", "nafta", "suprema", "optimo", "max s10", "s10")
    return any(token in normalized for token in accepted_tokens)


def _parse_number(value):
    s = str(value if value is not None else "").strip()
    if s == "":
        return 0.0
    if "," in s and "." in s:
        if s.find(",") < s.find("."):
            s = s.replace(",", "")
        else:
            s = s.replace(".", "").replace(",", ".")
    else:
        if "," in s:
            if s.count(",") >= 1:
                parts = s.split(",")
                if len(parts) > 2 or (len(parts) == 2 and len(parts[1]) == 3 and parts[0].replace("-", "").isdigit() and parts[1].isdigit()):
                    s = s.replace(",", "")
                else:
                    s = s.replace(",", ".")
        elif "." in s:
            parts = s.split(".")
            if len(parts) > 2 or (len(parts) == 2 and len(parts[1]) == 3 and parts[0].replace("-", "").isdigit() and parts[1].isdigit()):
                s = s.replace(".", "")
    return float(s)


def _parse_int(value):
    s = str(value if value is not None else "").strip()
    if s == "":
        return 0
    s = s.replace(",", "")
    if "." in s:
        raise ValueError("Unidades debe ser un entero.")
    return int(s)


def _parse_bool(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "t", "si", "sí", "y", "yes"}:
        return True
    if text in {"0", "false", "f", "no", "n"}:
        return False
    return default


def _iso_week_parts(value: date):
    iso_year, iso_week, _ = value.isocalendar()
    return iso_week, iso_year


def _shift_iso_week(week: int, year: int, offset: int):
    base = date.fromisocalendar(int(year), int(week), 1)
    shifted = base + timedelta(days=offset * 7)
    return _iso_week_parts(shifted)


def _shift_month(month: int, year: int, offset: int):
    total = (int(year) * 12 + (int(month) - 1)) + int(offset)
    shifted_year = total // 12
    shifted_month = total % 12 + 1
    return shifted_month, shifted_year


def _fmt_int(value):
    try:
        return f"{int(value):,}".replace(",", ".")
    except Exception:
        return str(value)


def _fmt_float(value, dec=2):
    try:
        txt = f"{float(value):,.{dec}f}"
        return txt.replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(value)


def _get_sucursal(slug: str):
    sucursal = SUCURSALES.get((slug or "").strip().lower())
    if not sucursal:
        raise ValueError("Sucursal no habilitada.")
    return sucursal


def _recepcion_slug_from_path(path: str):
    prefix = "/api/recepcion/"
    if not path.startswith(prefix):
        return None, ""
    rest = path[len(prefix):].strip("/")
    if not rest:
        return None, ""
    parts = rest.split("/")
    return parts[0], "/".join(parts[1:])


class DashboardRepository:
    """Consultas para el tablero web y la recepcion controlada."""

    def __init__(self, db_url: str):
        self.db_url = db_url
        self._schema_ready = False

    def _connect(self, readonly=True):
        self._ensure_schema()
        conn = psycopg2.connect(self.db_url, connect_timeout=5)
        conn.set_session(readonly=readonly, autocommit=readonly)
        return conn

    def _ensure_schema(self):
        if self._schema_ready:
            return
        with psycopg2.connect(self.db_url, connect_timeout=5) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    ALTER TABLE lotes
                    ADD COLUMN IF NOT EXISTS cerrado BOOLEAN NOT NULL DEFAULT FALSE
                    """
                )
                cur.execute(
                    """
                    ALTER TABLE lotes
                    ADD COLUMN IF NOT EXISTS cantidad_vac INTEGER NOT NULL DEFAULT 0,
                    ADD COLUMN IF NOT EXISTS cantidad_tor INTEGER NOT NULL DEFAULT 0,
                    ADD COLUMN IF NOT EXISTS cantidad_nov INTEGER NOT NULL DEFAULT 0,
                    ADD COLUMN IF NOT EXISTS cantidad_vaq INTEGER NOT NULL DEFAULT 0,
                    ADD COLUMN IF NOT EXISTS peso_promedio_vac NUMERIC(10, 2) NOT NULL DEFAULT 0,
                    ADD COLUMN IF NOT EXISTS peso_promedio_tor NUMERIC(10, 2) NOT NULL DEFAULT 0,
                    ADD COLUMN IF NOT EXISTS peso_promedio_nov NUMERIC(10, 2) NOT NULL DEFAULT 0,
                    ADD COLUMN IF NOT EXISTS peso_promedio_vaq NUMERIC(10, 2) NOT NULL DEFAULT 0
                    """
                )
                cur.execute(
                    """
                    ALTER TABLE cargas_combustible
                    ADD COLUMN IF NOT EXISTS eliminado_en TIMESTAMP NULL
                    """
                )
                cur.execute(
                    """
                    ALTER TABLE cargas_combustible
                    ADD COLUMN IF NOT EXISTS eliminado_por TEXT NULL
                    """
                )
                cur.execute(
                    """
                    ALTER TABLE cargas_combustible
                    ADD COLUMN IF NOT EXISTS motivo_eliminacion TEXT NULL
                    """
                )
                cur.execute(
                    """
                    ALTER TABLE gastos_flota
                    ADD COLUMN IF NOT EXISTS eliminado_en TIMESTAMP NULL
                    """
                )
                cur.execute(
                    """
                    ALTER TABLE gastos_flota
                    ADD COLUMN IF NOT EXISTS eliminado_por TEXT NULL
                    """
                )
                cur.execute(
                    """
                    ALTER TABLE gastos_flota
                    ADD COLUMN IF NOT EXISTS motivo_eliminacion TEXT NULL
                    """
                )
            conn.commit()
        self._schema_ready = True

    def _resumen_lotes_cte(self):
        return """
            SELECT
                L.id,
                L.lote,
                TRIM(L.empresa) AS empresa,
                L.fecha,
                COALESCE(L.cerrado, false) AS cerrado,
                L.cantidad AS cantcompra,
                COALESCE(F.faenado, 0) AS faenado,
                COALESCE(D.cabs_total, 0) AS distribuido,
                COALESCE(D.kg_total, 0.0) AS kg,
                L.monto,
                CASE
                    WHEN COALESCE(D.kg_total, 0) > 0
                        THEN ROUND((L.monto / D.kg_total)::numeric, 2)
                    ELSE 0
                END AS costokg,
                CASE
                    WHEN COALESCE(F.faenado, 0) > 0
                        THEN ROUND((COALESCE(D.cabs_total, 0)::numeric / F.faenado) * 100, 2)
                    ELSE 0
                END AS pct_distribuido,
                CASE
                    WHEN COALESCE(F.faenado, 0) > 0
                        THEN ROUND(100 - ((COALESCE(D.cabs_total, 0)::numeric / F.faenado) * 100), 2)
                    ELSE 100
                END AS pct_restante,
                COALESCE(L.peso_compra_kg, 0.0) AS kgcompra,
                COALESCE(L.cantidad_tor, 0)::int AS cantidad_tor,
                COALESCE(L.cantidad_nov, 0)::int AS cantidad_nov,
                COALESCE(L.cantidad_vac, 0)::int AS cantidad_vac,
                COALESCE(L.cantidad_vaq, 0)::int AS cantidad_vaq,
                COALESCE(L.peso_promedio_tor, 0.0) AS peso_promedio_tor,
                COALESCE(L.peso_promedio_nov, 0.0) AS peso_promedio_nov,
                COALESCE(L.peso_promedio_vac, 0.0) AS peso_promedio_vac,
                COALESCE(L.peso_promedio_vaq, 0.0) AS peso_promedio_vaq,
                CASE
                    WHEN COALESCE(L.peso_compra_kg, 0) > 0
                        THEN ROUND((COALESCE(D.kg_total, 0)::numeric / L.peso_compra_kg) * 100, 2)
                    ELSE 0
                END AS rend_pct
            FROM lotes L
            LEFT JOIN (
                SELECT lote_id, SUM(cantidad) AS faenado
                FROM faenas
                GROUP BY lote_id
            ) F ON F.lote_id = L.id
            LEFT JOIN (
                SELECT lote_id, SUM(kg) AS kg_total, SUM(cabezas) AS cabs_total
                FROM distribuciones
                GROUP BY lote_id
            ) D ON D.lote_id = L.id
        """

    def _menudencias_union_sql(self):
        return """
            SELECT 'Aregua' AS sucursal, fecha, producto, kg, unidades
            FROM menudencias_aregua
            UNION ALL
            SELECT 'Luque' AS sucursal, fecha, producto, kg, unidades
            FROM menudencias_luque
            UNION ALL
            SELECT 'Itaugua' AS sucursal, fecha, producto, kg, unidades
            FROM menudencias_itaugua
            UNION ALL
            SELECT sucursal, fecha, producto, kg, unidades
            FROM menudencias
            WHERE legacy_id IS NULL
        """

    def _costo_kg_default_ultimos_completados(self, cur):
        cur.execute(
            f"""
            WITH resumen_lotes AS ({self._resumen_lotes_cte()}),
            ultimos_completados AS (
                SELECT id, fecha, lote, monto, kg
                FROM resumen_lotes
                WHERE COALESCE(faenado, 0) > 0
                  AND COALESCE(distribuido, 0) = COALESCE(faenado, 0)
                  AND COALESCE(rend_pct, 0) > 45
                ORDER BY fecha DESC, lote
                LIMIT 4
            )
            SELECT CASE
                       WHEN COALESCE(SUM(kg), 0) > 0
                           THEN ROUND((SUM(monto) / SUM(kg))::numeric, 2)
                       ELSE 0
                   END AS costo_kg_promedio
            FROM ultimos_completados
            """
        )
        row = cur.fetchone() or {}
        return float(row["costo_kg_promedio"] or 0)

    def get_dashboard(self, desde=None, hasta=None):
        filters = []
        params = []
        if desde:
            filters.append("fecha >= %s")
            params.append(desde)
        if hasta:
            filters.append("fecha <= %s")
            params.append(hasta)
        where = "WHERE " + " AND ".join(filters) if filters else ""

        with self._connect() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """
                    WITH compras AS (
                        SELECT
                            COUNT(*)::int AS lotes,
                            COALESCE(SUM(cantidad), 0)::int AS reces_compradas,
                            COALESCE(SUM(peso_compra_kg), 0)::numeric AS kg_compra,
                            COALESCE(SUM(monto), 0)::numeric AS monto_total
                        FROM lotes
                        WHERE (%s::date IS NULL OR fecha >= %s::date)
                          AND (%s::date IS NULL OR fecha <= %s::date)
                    ),
                    faena AS (
                        SELECT COALESCE(SUM(cantidad), 0)::int AS reces_faenadas
                        FROM faenas
                        WHERE (%s::date IS NULL OR fecha >= %s::date)
                          AND (%s::date IS NULL OR fecha <= %s::date)
                    ),
                    distrib AS (
                        SELECT
                            COALESCE(SUM(cabezas), 0)::int AS reces_distribuidas,
                            COALESCE(SUM(kg), 0)::numeric AS kg_distribuidos
                        FROM distribuciones
                        WHERE (%s::date IS NULL OR fecha >= %s::date)
                          AND (%s::date IS NULL OR fecha <= %s::date)
                    )
                    SELECT
                        compras.lotes,
                        compras.reces_compradas,
                        faena.reces_faenadas,
                        distrib.reces_distribuidas,
                        distrib.kg_distribuidos,
                        compras.kg_compra,
                        compras.monto_total,
                        CASE WHEN distrib.kg_distribuidos > 0
                            THEN ROUND((compras.monto_total / distrib.kg_distribuidos)::numeric, 2)
                            ELSE 0
                        END AS costo_kg_promedio,
                        CASE WHEN faena.reces_faenadas > 0
                            THEN ROUND((distrib.reces_distribuidas::numeric / faena.reces_faenadas) * 100, 2)
                            ELSE 0
                        END AS pct_distribuido,
                        CASE WHEN compras.kg_compra > 0
                            THEN ROUND((distrib.kg_distribuidos::numeric / compras.kg_compra) * 100, 2)
                            ELSE 0
                        END AS rendimiento_pct
                    FROM compras
                    CROSS JOIN faena
                    CROSS JOIN distrib
                    """,
                    (desde, desde, hasta, hasta, desde, desde, hasta, hasta, desde, desde, hasta, hasta),
                )
                resumen = cur.fetchone() or {}
                resumen["costo_kg_promedio"] = self._costo_kg_default_ultimos_completados(cur)

                cur.execute(
                    """
                    WITH saldos AS (
                        SELECT L.id,
                               L.cantidad,
                               COALESCE(SUM(F.cantidad), 0) AS faenado,
                               L.cantidad - COALESCE(SUM(F.cantidad), 0) AS restante
                        FROM lotes L
                        LEFT JOIN faenas F ON F.lote_id = L.id
                        GROUP BY L.id, L.cantidad
                    ),
                    camara AS (
                        SELECT L.id,
                               COALESCE(SUM(F.cantidad), 0) AS faenado,
                               (
                                   SELECT COALESCE(SUM(D.cabezas), 0)
                                   FROM distribuciones D
                                   WHERE D.lote_id = L.id
                               ) AS distribuidas
                        FROM lotes L
                        LEFT JOIN faenas F ON F.lote_id = L.id
                        GROUP BY L.id
                        HAVING COALESCE(SUM(F.cantidad), 0) > 0
                    )
                    SELECT
                        COALESCE((SELECT SUM(GREATEST(faenado - distribuidas, 0)) FROM camara), 0)::int AS reses_camara,
                        COALESCE((SELECT SUM(GREATEST(restante, 0)) FROM saldos), 0)::int AS reses_sin_faenar,
                        COALESCE((SELECT COUNT(*) FROM saldos WHERE restante > 0), 0)::int AS lotes_pendientes
                    """
                )
                resumen_operativo = cur.fetchone() or {}
                resumen.update(resumen_operativo)

                cur.execute(
                    f"""
                    WITH resumen_lotes AS ({self._resumen_lotes_cte()})
                    SELECT empresa,
                           COALESCE(SUM(cantcompra), 0)::int AS reces,
                           COALESCE(SUM(kgcompra), 0)::numeric AS kg_compra,
                           COALESCE(SUM(monto), 0)::numeric AS monto
                    FROM resumen_lotes
                    {where}
                    GROUP BY empresa
                    ORDER BY reces DESC, empresa
                    """,
                    params,
                )
                compras_por_empresa = cur.fetchall()

                cur.execute(
                    """
                    SELECT local,
                           COALESCE(SUM(cabezas), 0)::int AS reces,
                           COALESCE(SUM(kg), 0)::numeric AS kg
                    FROM distribuciones
                    WHERE (%s::date IS NULL OR fecha >= %s::date)
                      AND (%s::date IS NULL OR fecha <= %s::date)
                    GROUP BY local
                    ORDER BY kg DESC, local
                    """,
                    (desde, desde, hasta, hasta),
                )
                distribuciones_por_local = cur.fetchall()

                cur.execute(
                    f"""
                    SELECT sucursal,
                           COALESCE(SUM(kg), 0)::numeric AS kg,
                           COALESCE(SUM(unidades), 0)::int AS unidades
                    FROM ({self._menudencias_union_sql()}) m
                    WHERE (%s::date IS NULL OR fecha >= %s::date)
                      AND (%s::date IS NULL OR fecha <= %s::date)
                    GROUP BY sucursal
                    ORDER BY kg DESC, sucursal
                    """,
                    (desde, desde, hasta, hasta),
                )
                menudencias_por_sucursal = cur.fetchall()

                cur.execute(
                    f"""
                    SELECT producto,
                           COALESCE(SUM(kg), 0)::numeric AS kg,
                           COALESCE(SUM(unidades), 0)::int AS unidades,
                           CASE WHEN COALESCE(SUM(unidades), 0) > 0
                               THEN ROUND((SUM(kg)::numeric / SUM(unidades)), 3)
                               ELSE 0
                           END AS kg_por_unidad
                    FROM ({self._menudencias_union_sql()}) m
                    WHERE (%s::date IS NULL OR fecha >= %s::date)
                      AND (%s::date IS NULL OR fecha <= %s::date)
                      AND TRIM(COALESCE(producto, '')) <> ''
                    GROUP BY producto
                    HAVING COALESCE(SUM(kg), 0) <> 0 OR COALESCE(SUM(unidades), 0) <> 0
                    ORDER BY kg DESC, unidades DESC, producto
                    LIMIT 12
                    """,
                    (desde, desde, hasta, hasta),
                )
                top_menudencias = cur.fetchall()

                cur.execute(
                    f"""
                    SELECT producto,
                           COALESCE(SUM(kg), 0)::numeric AS kg_total,
                           COALESCE(SUM(unidades), 0)::int AS unidades_total,
                           COALESCE(SUM(kg) FILTER (WHERE sucursal = 'Aregua'), 0)::numeric AS aregua_kg,
                           COALESCE(SUM(unidades) FILTER (WHERE sucursal = 'Aregua'), 0)::int AS aregua_unidades,
                           COALESCE(SUM(kg) FILTER (WHERE sucursal = 'Luque'), 0)::numeric AS luque_kg,
                           COALESCE(SUM(unidades) FILTER (WHERE sucursal = 'Luque'), 0)::int AS luque_unidades,
                           COALESCE(SUM(kg) FILTER (WHERE sucursal = 'Itaugua'), 0)::numeric AS itaugua_kg,
                           COALESCE(SUM(unidades) FILTER (WHERE sucursal = 'Itaugua'), 0)::int AS itaugua_unidades
                    FROM ({self._menudencias_union_sql()}) m
                    WHERE (%s::date IS NULL OR fecha >= %s::date)
                      AND (%s::date IS NULL OR fecha <= %s::date)
                      AND TRIM(COALESCE(producto, '')) <> ''
                    GROUP BY producto
                    HAVING COALESCE(SUM(kg), 0) <> 0 OR COALESCE(SUM(unidades), 0) <> 0
                    ORDER BY kg_total DESC, unidades_total DESC, producto
                    LIMIT 20
                    """,
                    (desde, desde, hasta, hasta),
                )
                menudencias_por_producto_sucursal = cur.fetchall()

                cur.execute(
                    f"""
                    WITH resumen_lotes AS ({self._resumen_lotes_cte()})
                    SELECT id,
                           lote,
                           empresa,
                           fecha,
                           cantcompra,
                           faenado,
                           distribuido,
                           kg,
                           kgcompra,
                           monto,
                           costokg,
                           pct_distribuido,
                           pct_restante,
                           rend_pct
                    FROM resumen_lotes
                    {where}
                    ORDER BY fecha DESC, lote
                    LIMIT 80
                    """,
                    params,
                )
                lotes = cur.fetchall()

        return {
            "resumen": resumen,
            "comprasPorEmpresa": compras_por_empresa,
            "distribucionesPorLocal": distribuciones_por_local,
            "menudenciasPorSucursal": menudencias_por_sucursal,
            "topMenudencias": top_menudencias,
            "menudenciasPorProductoSucursal": menudencias_por_producto_sucursal,
            "lotes": lotes,
        }

    def build_menudencias_pdf(self, desde=None, hasta=None, generated_by=None):
        if SimpleDocTemplate is None:
            raise RuntimeError("ReportLab no esta instalado. Instale reportlab para generar PDF.")
        if not desde or not hasta:
            today = date.today()
            monday_this_week = today - timedelta(days=today.weekday())
            desde = monday_this_week - timedelta(days=7)
            hasta = desde + timedelta(days=5)

        with self._connect(readonly=True) as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    f"""
                    SELECT sucursal,
                           COALESCE(SUM(kg), 0)::numeric AS kg,
                           COALESCE(SUM(unidades), 0)::int AS unidades,
                           COUNT(*)::int AS filas
                    FROM ({self._menudencias_union_sql()}) m
                    WHERE fecha >= %s AND fecha <= %s
                    GROUP BY sucursal
                    ORDER BY kg DESC, sucursal
                    """,
                    (desde, hasta),
                )
                por_sucursal = cur.fetchall()

                cur.execute(
                    f"""
                    SELECT producto,
                           COALESCE(SUM(kg), 0)::numeric AS kg,
                           COALESCE(SUM(unidades), 0)::int AS unidades,
                           CASE WHEN COALESCE(SUM(unidades), 0) > 0
                               THEN ROUND((SUM(kg)::numeric / SUM(unidades)), 3)
                               ELSE 0
                           END AS kg_por_unidad
                    FROM ({self._menudencias_union_sql()}) m
                    WHERE fecha >= %s AND fecha <= %s
                      AND TRIM(COALESCE(producto, '')) <> ''
                    GROUP BY producto
                    HAVING COALESCE(SUM(kg), 0) <> 0 OR COALESCE(SUM(unidades), 0) <> 0
                    ORDER BY kg DESC, unidades DESC, producto
                    """,
                    (desde, hasta),
                )
                general = cur.fetchall()

                cur.execute(
                    f"""
                    SELECT sucursal,
                           producto,
                           COALESCE(SUM(kg), 0)::numeric AS kg,
                           COALESCE(SUM(unidades), 0)::int AS unidades,
                           CASE WHEN COALESCE(SUM(unidades), 0) > 0
                               THEN ROUND((SUM(kg)::numeric / SUM(unidades)), 3)
                               ELSE 0
                           END AS kg_por_unidad
                    FROM ({self._menudencias_union_sql()}) m
                    WHERE fecha >= %s AND fecha <= %s
                      AND TRIM(COALESCE(producto, '')) <> ''
                    GROUP BY sucursal, producto
                    HAVING COALESCE(SUM(kg), 0) <> 0 OR COALESCE(SUM(unidades), 0) <> 0
                    ORDER BY sucursal, kg DESC, unidades DESC, producto
                    """,
                    (desde, hasta),
                )
                detalle = cur.fetchall()

        total_kg = sum(float(row.get("kg") or 0) for row in general)
        total_unidades = sum(int(row.get("unidades") or 0) for row in general)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("MenudenciasTitle", parent=styles["Title"], fontName="Courier-Bold", fontSize=14, leading=16, spaceAfter=4)
        body_style = ParagraphStyle("MenudenciasBody", parent=styles["Normal"], fontName="Courier", fontSize=8, leading=9, spaceBefore=0, spaceAfter=2)
        heading_style = ParagraphStyle("MenudenciasHeading", parent=styles["Heading3"], fontName="Courier-Bold", fontSize=9, leading=10, spaceBefore=6, spaceAfter=4)
        kpi_value_style = ParagraphStyle("MenudenciasKpiValue", parent=styles["Normal"], fontName="Courier-Bold", fontSize=14, leading=15, alignment=1)

        kpi_data = [[
            self._pdf_wrap_cell_typewriter("Kg total", align="CENTER", font_size=8, leading=9),
            self._pdf_wrap_cell_typewriter("Unidades", align="CENTER", font_size=8, leading=9),
            self._pdf_wrap_cell_typewriter("Productos", align="CENTER", font_size=8, leading=9),
            self._pdf_wrap_cell_typewriter("Sucursales", align="CENTER", font_size=8, leading=9),
        ], [
            Paragraph(_fmt_float(total_kg, 2), kpi_value_style),
            Paragraph(_fmt_int(total_unidades), kpi_value_style),
            Paragraph(_fmt_int(len(general)), kpi_value_style),
            Paragraph(_fmt_int(len(por_sucursal)), kpi_value_style),
        ]]
        story = [
            Paragraph("<b>Resumen de menudencias</b>", title_style),
            Paragraph(
                f"Periodo: <b>{desde}</b> al <b>{hasta}</b> | Generado: <b>{datetime.now().strftime('%Y-%m-%d %H:%M')}</b> | Usuario: <b>{escape(str(generated_by or 'Sistema'))}</b>",
                body_style,
            ),
            Spacer(0, 6),
            self._build_table_compact_typewriter(kpi_data, col_widths=[150, 150, 120, 120], header_fill=colors.HexColor("#FFF3CD")),
            Spacer(0, 8),
            Paragraph("<b>General por producto</b>", heading_style),
        ]

        data_general = [["Producto", "Kg", "Unidades", "Kg/unidad"]]
        for row in general:
            data_general.append([
                self._pdf_wrap_cell_typewriter(row.get("producto") or "", align="LEFT", font_size=7, leading=8),
                _fmt_float(row.get("kg"), 2),
                _fmt_int(row.get("unidades")),
                _fmt_float(row.get("kg_por_unidad"), 3),
            ])
        if len(data_general) == 1:
            data_general.append(["Sin menudencias para el periodo", "0,00", "0", "0,000"])
        story.append(self._build_table_compact_typewriter(data_general, col_widths=[340, 90, 90, 90]))
        story.append(Spacer(0, 8))

        story.append(Paragraph("<b>Totales por sucursal</b>", heading_style))
        data_sucursal = [["Sucursal", "Kg", "Unidades", "Registros"]]
        for row in por_sucursal:
            data_sucursal.append([row.get("sucursal") or "", _fmt_float(row.get("kg"), 2), _fmt_int(row.get("unidades")), _fmt_int(row.get("filas"))])
        if len(data_sucursal) == 1:
            data_sucursal.append(["Sin datos", "0,00", "0", "0"])
        story.append(self._build_table_compact_typewriter(data_sucursal, col_widths=[180, 120, 120, 100]))
        story.append(Spacer(0, 8))

        for sucursal in ("Aregua", "Luque", "Itaugua"):
            rows = [row for row in detalle if row.get("sucursal") == sucursal]
            story.append(Paragraph(f"<b>{sucursal}</b>", heading_style))
            data_detalle = [["Producto", "Kg", "Unidades", "Kg/unidad"]]
            for row in rows:
                data_detalle.append([
                    self._pdf_wrap_cell_typewriter(row.get("producto") or "", align="LEFT", font_size=7, leading=8),
                    _fmt_float(row.get("kg"), 2),
                    _fmt_int(row.get("unidades")),
                    _fmt_float(row.get("kg_por_unidad"), 3),
                ])
            if len(data_detalle) == 1:
                data_detalle.append(["Sin menudencias para el periodo", "0,00", "0", "0,000"])
            story.append(self._build_table_compact_typewriter(data_detalle, col_widths=[340, 90, 90, 90]))
            story.append(Spacer(0, 8))

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
        doc.build(story)
        return buffer.getvalue(), f"Resumen_Menudencias_{desde}_{hasta}.pdf"

    def get_recepcion(self, slug, fecha=None):
        sucursal = _get_sucursal(slug)
        fecha = fecha or date.today()
        desde = date.today() - timedelta(days=1)
        with self._connect(readonly=True) as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """
                    SELECT d.id,
                           d.lote_id,
                           l.lote,
                           l.empresa AS origen,
                           d.fecha,
                           d.kg,
                           d.cabezas,
                           COALESCE(d.nota, '') AS nota,
                           d.local,
                           COALESCE(d.diferencia_kg, 0) AS diferencia_kg
                    FROM distribuciones d
                    JOIN lotes l ON l.id = d.lote_id
                    WHERE d.local = %s AND d.fecha >= %s
                    ORDER BY d.fecha DESC, d.id DESC
                    """,
                    (sucursal["local"], desde),
                )
                distribuciones = cur.fetchall()

                cur.execute(
                    """
                    SELECT id, fecha, producto, kg, unidades
                    FROM menudencias
                    WHERE sucursal = %s
                      AND fecha = %s
                    ORDER BY producto, id
                    """,
                    (sucursal["nombre"], fecha),
                )
                menudencias = cur.fetchall()

        return {
            "fecha": fecha,
            "sucursal": sucursal["nombre"],
            "local": sucursal["local"],
            "distribuciones": distribuciones,
            "menudencias": menudencias,
        }

    def build_recepcion_pdf(self, slug, fecha=None, generated_by=None, user_role=None, allowed_scope=None):
        if SimpleDocTemplate is None:
            raise RuntimeError("ReportLab no esta instalado. Instale reportlab para generar PDF.")

        recepcion = self.get_recepcion(slug, fecha=fecha)
        distribuciones = recepcion.get("distribuciones") or []
        menudencias = recepcion.get("menudencias") or []
        total_kg_distribuido = sum(float(row.get("kg") or 0) for row in distribuciones)
        total_cabezas = sum(int(row.get("cabezas") or 0) for row in distribuciones)
        total_dif_kg = sum(float(row.get("diferencia_kg") or 0) for row in distribuciones)
        total_kg_menudencias = sum(float(row.get("kg") or 0) for row in menudencias)
        total_unidades_menudencias = sum(int(row.get("unidades") or 0) for row in menudencias)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "RecepcionTypeTitle",
            parent=styles["Title"],
            fontName="Courier-Bold",
            fontSize=14,
            leading=16,
            spaceAfter=4,
        )
        heading_style = ParagraphStyle(
            "RecepcionTypeHeading",
            parent=styles["Heading3"],
            fontName="Courier-Bold",
            fontSize=9,
            leading=10,
            spaceBefore=2,
            spaceAfter=4,
        )
        body_style = ParagraphStyle(
            "RecepcionTypeBody",
            parent=styles["Normal"],
            fontName="Courier",
            fontSize=8,
            leading=9,
            spaceBefore=0,
            spaceAfter=2,
        )
        kpi_value_style = ParagraphStyle(
            "RecepcionTypeKpiValue",
            parent=styles["Normal"],
            fontName="Courier-Bold",
            fontSize=14,
            leading=15,
            alignment=1,
            spaceBefore=0,
            spaceAfter=0,
        )

        kpi_data = [[
            self._pdf_wrap_cell_typewriter("Distribuciones", align="CENTER", font_size=8, leading=9),
            self._pdf_wrap_cell_typewriter("Kg distribuidos", align="CENTER", font_size=8, leading=9),
            self._pdf_wrap_cell_typewriter("Reces", align="CENTER", font_size=8, leading=9),
            self._pdf_wrap_cell_typewriter("Kg menudencias", align="CENTER", font_size=8, leading=9),
        ], [
            Paragraph(_fmt_int(len(distribuciones)), kpi_value_style),
            Paragraph(_fmt_float(total_kg_distribuido, 2), kpi_value_style),
            Paragraph(_fmt_int(total_cabezas), kpi_value_style),
            Paragraph(_fmt_float(total_kg_menudencias, 2), kpi_value_style),
        ]]

        story = [
            Paragraph("<b>Reporte de recepcion</b>", title_style),
            Paragraph(
                (
                    f"Sucursal: <b>{escape(str(recepcion.get('sucursal') or ''))}</b> | "
                    f"Fecha menudencias: <b>{escape(str(recepcion.get('fecha') or ''))}</b> | "
                    f"Generado: <b>{datetime.now().strftime('%Y-%m-%d %H:%M')}</b>"
                ),
                body_style,
            ),
            Paragraph(
                (
                    f"Operativa: <b>{escape(str(recepcion.get('local') or ''))}</b> | "
                    f"Usuario: <b>{escape(str(generated_by or 'Sistema'))}</b> | "
                    f"Rol: <b>{escape(str(user_role or '-'))}</b> | "
                    f"Sucursal habilitada: <b>{escape(str(allowed_scope or 'todas'))}</b>"
                ),
                body_style,
            ),
            Spacer(0, 6),
            self._build_table_compact_typewriter(kpi_data, col_widths=[130, 150, 100, 150], header_fill=colors.HexColor("#FFF3CD")),
            Spacer(0, 8),
        ]

        data_dist = [["Fecha", "Lote", "Origen", "Kg", "Cabezas", "Dif. Kg", "Nota"]]
        for row in distribuciones:
            data_dist.append([
                str(row.get("fecha") or ""),
                str(row.get("lote") or ""),
                self._pdf_wrap_cell_typewriter(row.get("origen") or "", align="LEFT", font_size=7, leading=8),
                _fmt_float(row.get("kg"), 2),
                _fmt_int(row.get("cabezas")),
                _fmt_float(row.get("diferencia_kg"), 2),
                self._pdf_wrap_cell_typewriter(row.get("nota") or "", align="LEFT", font_size=7, leading=8),
            ])
        if len(data_dist) == 1:
            data_dist.append(["-", "-", "-", "0,00", "0", "0,00", "Sin distribuciones para el filtro actual"])

        story.append(Paragraph(f"<b>Distribuciones</b> | Diferencia acumulada: <b>{_fmt_float(total_dif_kg, 2)} kg</b>", heading_style))
        story.append(self._build_table_compact_typewriter(data_dist, col_widths=[60, 60, 120, 60, 55, 60, 250]))
        story.append(Spacer(0, 8))

        data_men = [["Producto", "Kg", "Unidades"]]
        for row in menudencias:
            data_men.append([
                self._pdf_wrap_cell_typewriter(row.get("producto") or "", align="LEFT", font_size=7, leading=8),
                _fmt_float(row.get("kg"), 2),
                _fmt_int(row.get("unidades")),
            ])
        if len(data_men) == 1:
            data_men.append(["Sin menudencias cargadas para la fecha seleccionada", "0,00", "0"])

        story.append(
            Paragraph(
                f"<b>Menudencias</b> | Total kg: <b>{_fmt_float(total_kg_menudencias, 2)}</b> | Total unidades: <b>{_fmt_int(total_unidades_menudencias)}</b>",
                heading_style,
            )
        )
        story.append(self._build_table_compact_typewriter(data_men, col_widths=[360, 90, 90]))

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=18,
            rightMargin=18,
            topMargin=18,
            bottomMargin=18,
        )
        doc.build(story)
        fecha_txt = str(recepcion.get("fecha") or date.today().isoformat())
        sucursal_txt = str(recepcion.get("sucursal") or slug).replace(" ", "_")
        return buffer.getvalue(), f"Reporte_Recepcion_{sucursal_txt}_{fecha_txt}.pdf"

    def update_recepcion_distribucion(self, slug, payload):
        sucursal = _get_sucursal(slug)
        distrib_id = int(payload.get("id"))
        nuevo_kg = _parse_number(payload.get("kg"))
        if nuevo_kg <= 0:
            raise ValueError("El kg recibido debe ser mayor a 0.")

        faltante = _parse_number(payload.get("faltante_kg") or 0)
        sobrante = _parse_number(payload.get("sobrante_kg") or 0)
        if faltante < 0 or sobrante < 0:
            raise ValueError("Faltante y sobrante no pueden ser negativos.")
        if faltante > 0 and sobrante > 0:
            raise ValueError("Completa solo faltante o sobrante, no ambos.")

        dif_kg = 0.0
        if faltante > 0:
            dif_kg = -abs(faltante)
        elif sobrante > 0:
            dif_kg = abs(sobrante)

        registrado_por = str(payload.get("registrado_por") or "").strip()
        if registrado_por:
            marca = f"Modificado por {registrado_por} el {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        else:
            marca = f"Actualizado por recepcion {sucursal['nombre']} el {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        with self._connect(readonly=False) as conn:
            try:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute(
                        """
                        SELECT d.id,
                               d.lote_id,
                               d.fecha,
                               d.local,
                               d.cabezas,
                               COALESCE(d.nota, '') AS nota
                        FROM distribuciones d
                        WHERE d.id = %s AND d.local = %s
                        """,
                        (distrib_id, sucursal["local"]),
                    )
                    row = cur.fetchone()
                    if not row:
                        raise ValueError(f"Distribucion de {sucursal['nombre']} no encontrada.")

                    nota_base = str(payload.get("nota") if payload.get("nota") is not None else row["nota"]).strip()
                    base_prefix = f"Recepcionado en Cacique {sucursal['nombre']}"
                    cleaned = nota_base
                    if cleaned.lower().startswith(base_prefix.lower()):
                        cleaned = cleaned[len(base_prefix):].lstrip()
                        if cleaned.startswith("."):
                            cleaned = cleaned[1:].lstrip()
                    nueva_nota = base_prefix if not cleaned else f"{base_prefix}. {cleaned}"
                    nueva_nota = f"{nueva_nota} | {marca}"

                    cur.execute(
                        """
                        UPDATE distribuciones
                        SET kg = %s, nota = %s, diferencia_kg = %s
                        WHERE id = %s AND local = %s
                        RETURNING id, lote_id, fecha, local, kg, cabezas, nota, diferencia_kg
                        """,
                        (nuevo_kg, nueva_nota, dif_kg, distrib_id, sucursal["local"]),
                    )
                    updated = cur.fetchone()
                conn.commit()
                return updated
            except Exception:
                conn.rollback()
                raise

    def add_menudencia(self, slug, payload):
        sucursal = _get_sucursal(slug)
        fecha = _parse_date(payload.get("fecha")) or date.today()
        producto = str(payload.get("producto") or "").strip()
        if not producto:
            raise ValueError("Ingresar el nombre del producto.")
        kg = _parse_number(payload.get("kg") or 0)
        unidades = _parse_int(payload.get("unidades") or 0)
        if kg < 0 or unidades < 0:
            raise ValueError("Kg y unidades no pueden ser negativos.")

        with self._connect(readonly=False) as conn:
            try:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    self._ensure_catalogo_producto(cur, producto)
                    cur.execute(
                        """
                        INSERT INTO menudencias(sucursal, fecha, producto, kg, unidades)
                        VALUES (%s, %s, %s, %s, %s)
                        RETURNING id, fecha, producto, kg, unidades
                        """,
                        (sucursal["nombre"], fecha, producto, kg, unidades),
                    )
                    row = cur.fetchone()
                conn.commit()
                return row
            except Exception:
                conn.rollback()
                raise

    def update_menudencia(self, slug, payload):
        sucursal = _get_sucursal(slug)
        men_id = int(payload.get("id"))
        producto = str(payload.get("producto") or "").strip()
        if not producto:
            raise ValueError("Producto no puede quedar vacio.")
        kg = _parse_number(payload.get("kg") or 0)
        unidades = _parse_int(payload.get("unidades") or 0)
        if kg < 0 or unidades < 0:
            raise ValueError("Kg y unidades no pueden ser negativos.")

        with self._connect(readonly=False) as conn:
            try:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    self._ensure_catalogo_producto(cur, producto)
                    cur.execute(
                        """
                        UPDATE menudencias
                        SET producto = %s, kg = %s, unidades = %s
                        WHERE id = %s
                          AND sucursal = %s
                        RETURNING id, fecha, producto, kg, unidades
                        """,
                        (producto, kg, unidades, men_id, sucursal["nombre"]),
                    )
                    row = cur.fetchone()
                    if not row:
                        raise ValueError("Menudencia no encontrada.")
                conn.commit()
                return row
            except Exception:
                conn.rollback()
                raise

    def delete_menudencia(self, slug, men_id):
        sucursal = _get_sucursal(slug)
        with self._connect(readonly=False) as conn:
            try:
                with conn.cursor() as cur:
                    cur.execute(
                        "DELETE FROM menudencias WHERE id = %s AND sucursal = %s",
                        (int(men_id), sucursal["nombre"]),
                    )
                    if cur.rowcount == 0:
                        raise ValueError("Menudencia no encontrada.")
                conn.commit()
            except Exception:
                conn.rollback()
                raise
        return {"ok": True}

    def _ensure_catalogo_producto(self, cur, producto):
        cur.execute(
            """
            SELECT 1
            FROM menudencias_catalogo
            WHERE LOWER(producto) = LOWER(%s)
            LIMIT 1
            """,
            (producto,),
        )
        if cur.fetchone() is None:
            cur.execute("INSERT INTO menudencias_catalogo(producto) VALUES (%s)", (producto,))

    def get_flota_catalogos(self, sucursal=None):
        filters = []
        params: list[Any] = []
        if sucursal:
            filters.append("LOWER(COALESCE(sucursal, '')) = %s")
            params.append(str(sucursal).strip().lower())
        where = f"WHERE {' AND '.join(filters)}" if filters else ""

        with self._connect(readonly=True) as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    f"""
                    SELECT id, codigo, chapa, nombre, marca, modelo, anho, tipo, sucursal, chofer, activo, creado_en
                    FROM vehiculos
                    {where}
                    ORDER BY activo DESC, nombre, chapa
                    """,
                    params,
                )
                vehiculos = cur.fetchall()

                cur.execute(
                    """
                    SELECT id, nombre, tipo, ruc, telefono, activo
                    FROM proveedores_flota
                    ORDER BY activo DESC, tipo, nombre
                    """
                )
                proveedores = cur.fetchall()

                cur.execute(
                    """
                    SELECT id, nombre, requiere_km, activo
                    FROM tipos_gasto_flota
                    ORDER BY nombre
                    """
                )
                tipos_gasto = cur.fetchall()

        return {
            "vehiculos": vehiculos,
            "proveedores": proveedores,
            "tiposGasto": tipos_gasto,
            "sucursales": [
                {"slug": slug, "nombre": data["nombre"], "local": data["local"]}
                for slug, data in SUCURSALES.items()
                if not sucursal or slug == str(sucursal).strip().lower()
            ],
        }

    def list_vehiculos(self, activo=None, sucursal=None):
        filters = []
        params: list[Any] = []
        if activo is not None:
            filters.append("activo = %s")
            params.append(bool(activo))
        if sucursal:
            filters.append("LOWER(COALESCE(sucursal, '')) = %s")
            params.append(str(sucursal).strip().lower())
        where = f"WHERE {' AND '.join(filters)}" if filters else ""

        with self._connect(readonly=True) as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    f"""
                    SELECT id, codigo, chapa, nombre, marca, modelo, anho, tipo, sucursal, chofer, activo, creado_en
                    FROM vehiculos
                    {where}
                    ORDER BY activo DESC, nombre, chapa
                    """,
                    params,
                )
                rows = cur.fetchall()
        return {"items": rows}

    def save_vehiculo(self, payload, sucursal_scope=None):
        vehiculo_id = payload.get("id")
        codigo = str(payload.get("codigo") or "").strip().upper() or None
        chapa = str(payload.get("chapa") or "").strip().upper() or None
        nombre = str(payload.get("nombre") or "").strip()
        marca = str(payload.get("marca") or "").strip()
        modelo = str(payload.get("modelo") or "").strip()
        tipo = str(payload.get("tipo") or "").strip()
        sucursal = str(payload.get("sucursal") or "").strip().lower() or None
        chofer = str(payload.get("chofer") or "").strip()
        activo = _parse_bool(payload.get("activo"), default=True)
        anho = payload.get("anho")
        anho = int(anho) if str(anho or "").strip() else None
        scoped_sucursal = str(sucursal_scope or "").strip().lower() or None

        if scoped_sucursal:
            sucursal = scoped_sucursal

        if not any([codigo, chapa, nombre, marca, modelo, tipo, chofer, anho]):
            raise ValueError("Debes completar al menos un dato del vehiculo.")
        if sucursal and sucursal not in SUCURSALES:
            raise ValueError("Sucursal invalida para el vehiculo.")

        with self._connect(readonly=False) as conn:
            try:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    if vehiculo_id:
                        if scoped_sucursal:
                            cur.execute("SELECT sucursal FROM vehiculos WHERE id = %s", (int(vehiculo_id),))
                            existing = cur.fetchone()
                            if not existing:
                                raise ValueError("Vehiculo no encontrado.")
                            if str(existing.get("sucursal") or "").strip().lower() != scoped_sucursal:
                                raise PermissionDenied("No tienes permisos para editar vehiculos de otra sucursal.")
                        cur.execute(
                            """
                            UPDATE vehiculos
                            SET codigo = %s,
                                chapa = %s,
                                nombre = %s,
                                marca = %s,
                                modelo = %s,
                                anho = %s,
                                tipo = %s,
                                sucursal = %s,
                                chofer = %s,
                                activo = %s
                            WHERE id = %s
                            RETURNING id, codigo, chapa, nombre, marca, modelo, anho, tipo, sucursal, chofer, activo, creado_en
                            """,
                            (codigo, chapa, nombre, marca, modelo, anho, tipo, sucursal, chofer, activo, int(vehiculo_id)),
                        )
                    else:
                        cur.execute(
                            """
                            INSERT INTO vehiculos(codigo, chapa, nombre, marca, modelo, anho, tipo, sucursal, chofer, activo)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            RETURNING id, codigo, chapa, nombre, marca, modelo, anho, tipo, sucursal, chofer, activo, creado_en
                            """,
                            (codigo, chapa, nombre, marca, modelo, anho, tipo, sucursal, chofer, activo),
                        )
                    row = cur.fetchone()
                    if not row:
                        raise ValueError("Vehiculo no encontrado.")
                conn.commit()
                return row
            except psycopg2.errors.UniqueViolation:
                conn.rollback()
                raise ValueError("Codigo o chapa ya registrados para otro vehiculo.")
            except Exception:
                conn.rollback()
                raise

    def save_proveedor_flota(self, payload):
        proveedor_id = payload.get("id")
        nombre = str(payload.get("nombre") or "").strip()
        tipo = str(payload.get("tipo") or "").strip().lower()
        ruc = str(payload.get("ruc") or "").strip()
        telefono = str(payload.get("telefono") or "").strip()
        activo = _parse_bool(payload.get("activo"), default=True)

        if not nombre:
            raise ValueError("El nombre del proveedor es obligatorio.")
        if tipo not in {"combustible", "taller", "otros"}:
            raise ValueError("Tipo de proveedor invalido.")

        with self._connect(readonly=False) as conn:
            try:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    if proveedor_id:
                        cur.execute(
                            """
                            UPDATE proveedores_flota
                            SET nombre = %s, tipo = %s, ruc = %s, telefono = %s, activo = %s
                            WHERE id = %s
                            RETURNING id, nombre, tipo, ruc, telefono, activo
                            """,
                            (nombre, tipo, ruc, telefono, activo, int(proveedor_id)),
                        )
                    else:
                        cur.execute(
                            """
                            INSERT INTO proveedores_flota(nombre, tipo, ruc, telefono, activo)
                            VALUES (%s, %s, %s, %s, %s)
                            RETURNING id, nombre, tipo, ruc, telefono, activo
                            """,
                            (nombre, tipo, ruc, telefono, activo),
                        )
                    row = cur.fetchone()
                    if not row:
                        raise ValueError("Proveedor no encontrado.")
                conn.commit()
                return row
            except psycopg2.errors.UniqueViolation:
                conn.rollback()
                raise ValueError("Ya existe un proveedor con ese nombre y tipo.")
            except Exception:
                conn.rollback()
                raise

    def list_cargas_combustible(self, desde=None, hasta=None, vehiculo_id=None, sucursal=None):
        filters = ["c.eliminado_en IS NULL"]
        params: list[Any] = []
        if desde:
            filters.append("c.fecha >= %s")
            params.append(desde)
        if hasta:
            filters.append("c.fecha <= %s")
            params.append(hasta)
        if vehiculo_id:
            filters.append("c.vehiculo_id = %s")
            params.append(int(vehiculo_id))
        if sucursal:
            filters.append("LOWER(COALESCE(v.sucursal, '')) = %s")
            params.append(str(sucursal).strip().lower())
        where = f"WHERE {' AND '.join(filters)}" if filters else ""

        with self._connect(readonly=True) as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    f"""
                    SELECT c.id,
                           c.fecha,
                           c.vehiculo_id,
                           v.codigo AS vehiculo_codigo,
                           v.nombre AS vehiculo_nombre,
                           v.chapa,
                           v.sucursal,
                           c.proveedor_id,
                           COALESCE(p.nombre, '') AS proveedor_nombre,
                           c.litros,
                           c.importe,
                           c.precio_litro,
                           COALESCE(c.tipo_combustible, '') AS tipo_combustible,
                           c.km_actual,
                           COALESCE(c.nro_factura, '') AS nro_factura,
                           COALESCE(c.observacion, '') AS observacion,
                           c.semana,
                           c.anho,
                           c.cargado_por,
                           c.creado_en,
                           c.eliminado_en,
                           COALESCE(c.eliminado_por, '') AS eliminado_por,
                           COALESCE(c.motivo_eliminacion, '') AS motivo_eliminacion
                    FROM cargas_combustible c
                    JOIN vehiculos v ON v.id = c.vehiculo_id
                    LEFT JOIN proveedores_flota p ON p.id = c.proveedor_id
                    {where}
                    ORDER BY c.fecha DESC, c.id DESC
                    LIMIT 250
                    """,
                    params,
                )
                rows = cur.fetchall()
        return {"items": rows}

    def save_carga_combustible(self, payload, cargado_por=None, sucursal_scope=None):
        carga_id = _parse_int(payload.get("id") or 0)
        fecha = _parse_date(payload.get("fecha"))
        vehiculo_id = int(payload.get("vehiculo_id") or 0)
        proveedor_id = payload.get("proveedor_id")
        proveedor_id = int(proveedor_id) if str(proveedor_id or "").strip() else None
        litros = _parse_number(payload.get("litros") or 0)
        importe = _parse_number(payload.get("importe") or 0)
        km_actual_raw = payload.get("km_actual")
        km_actual = _parse_number(km_actual_raw) if str(km_actual_raw or "").strip() else None
        tipo_combustible = str(payload.get("tipo_combustible") or "").strip()
        nro_factura = str(payload.get("nro_factura") or "").strip()
        observacion = str(payload.get("observacion") or "").strip()

        if not fecha:
            raise ValueError("La fecha es obligatoria.")
        if vehiculo_id <= 0:
            raise ValueError("Debes seleccionar un vehiculo.")
        if litros <= 0:
            raise ValueError("Los litros deben ser mayores a 0.")
        if importe <= 0:
            raise ValueError("El importe debe ser mayor a 0.")

        semana, anho = _iso_week_parts(fecha)
        precio_litro = round(importe / litros, 2)

        with self._connect(readonly=False) as conn:
            try:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    vehiculo = self._require_vehiculo_activo(cur, vehiculo_id)
                    self._validate_flota_sucursal_scope(vehiculo, sucursal_scope)
                    self._validate_proveedor_tipo(cur, proveedor_id, "combustible")
                    self._validate_km_flota(cur, vehiculo_id, km_actual)
                    self._validate_factura_combustible_unica(cur, proveedor_id, nro_factura, exclude_id=carga_id or None)
                    if carga_id:
                        cur.execute(
                            """
                            SELECT id
                            FROM cargas_combustible
                            WHERE id = %s AND eliminado_en IS NULL
                            """,
                            (carga_id,),
                        )
                        if not cur.fetchone():
                            raise ValueError("Carga de combustible no encontrada.")
                        cur.execute(
                            """
                            UPDATE cargas_combustible
                            SET vehiculo_id = %s,
                                fecha = %s,
                                proveedor_id = %s,
                                litros = %s,
                                importe = %s,
                                precio_litro = %s,
                                tipo_combustible = %s,
                                km_actual = %s,
                                nro_factura = %s,
                                observacion = %s,
                                semana = %s,
                                anho = %s
                            WHERE id = %s
                            RETURNING id, fecha, vehiculo_id, proveedor_id, litros, importe, precio_litro,
                                      tipo_combustible, km_actual, nro_factura, observacion, semana, anho, cargado_por, creado_en
                            """,
                            (
                                vehiculo_id,
                                fecha,
                                proveedor_id,
                                litros,
                                importe,
                                precio_litro,
                                tipo_combustible or None,
                                km_actual,
                                nro_factura,
                                observacion,
                                semana,
                                anho,
                                carga_id,
                            ),
                        )
                    else:
                        cur.execute(
                            """
                            INSERT INTO cargas_combustible(
                                vehiculo_id, fecha, proveedor_id, litros, importe, precio_litro,
                                tipo_combustible, km_actual, nro_factura, observacion, semana, anho, cargado_por
                            )
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            RETURNING id, fecha, vehiculo_id, proveedor_id, litros, importe, precio_litro,
                                      tipo_combustible, km_actual, nro_factura, observacion, semana, anho, cargado_por, creado_en
                            """,
                            (
                                vehiculo_id,
                                fecha,
                                proveedor_id,
                                litros,
                                importe,
                                precio_litro,
                                tipo_combustible or None,
                                km_actual,
                                nro_factura,
                                observacion,
                                semana,
                                anho,
                                str(cargado_por or "").strip() or None,
                            ),
                        )
                    row = cur.fetchone()
                    row["vehiculo_nombre"] = vehiculo["nombre"]
                    row["vehiculo_codigo"] = vehiculo["codigo"]
                    row["chapa"] = vehiculo["chapa"]
                    row["sucursal"] = vehiculo["sucursal"]
                conn.commit()
                return row
            except psycopg2.errors.UniqueViolation:
                conn.rollback()
                raise ValueError("La factura ya esta registrada en combustible.")
            except Exception:
                conn.rollback()
                raise

    def delete_carga_combustible(self, payload, eliminado_por=None, sucursal_scope=None):
        carga_id = _parse_int(payload.get("id") or 0)
        motivo = str(payload.get("motivo") or "").strip()

        if carga_id <= 0:
            raise ValueError("Falta el id de la carga.")
        if not motivo:
            raise ValueError("Debes indicar el motivo de eliminacion.")

        with self._connect(readonly=False) as conn:
            try:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute(
                        """
                        SELECT c.id,
                               c.eliminado_en,
                               c.vehiculo_id,
                               v.codigo,
                               v.nombre,
                               v.chapa,
                               v.sucursal
                        FROM cargas_combustible c
                        JOIN vehiculos v ON v.id = c.vehiculo_id
                        WHERE c.id = %s
                        """,
                        (carga_id,),
                    )
                    row = cur.fetchone()
                    if not row:
                        raise ValueError("La carga no existe.")
                    self._validate_flota_sucursal_scope(row, sucursal_scope)
                    if row.get("eliminado_en"):
                        raise ValueError("La carga ya fue eliminada.")
                    cur.execute(
                        """
                        UPDATE cargas_combustible
                        SET eliminado_en = NOW(),
                            eliminado_por = %s,
                            motivo_eliminacion = %s
                        WHERE id = %s
                        RETURNING id, eliminado_en, COALESCE(eliminado_por, '') AS eliminado_por, COALESCE(motivo_eliminacion, '') AS motivo_eliminacion
                        """,
                        (str(eliminado_por or "").strip() or None, motivo, carga_id),
                    )
                    deleted = cur.fetchone()
                conn.commit()
                return {"ok": True, "item": deleted}
            except Exception:
                conn.rollback()
                raise

    def import_cargas_combustible(self, payload, cargado_por=None, sucursal_scope=None):
        file_name = str(payload.get("file_name") or "").strip()
        file_content = str(payload.get("file_content") or "").strip()
        proveedor_id = payload.get("proveedor_id")
        proveedor_id = int(proveedor_id) if str(proveedor_id or "").strip() else None

        if not file_name or not file_content:
            raise ValueError("Debes adjuntar un archivo CSV o XLSX.")

        rows = self._read_combustible_import_rows(file_name, file_content)
        if not rows:
            raise ValueError("No se encontraron filas para importar.")

        inserted = 0
        skipped = 0
        errors = []
        missing_vehicles: dict[str, int] = {}
        seen_facturas: dict[str, int] = {}

        with self._connect(readonly=False) as conn:
            try:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    self._validate_proveedor_tipo(cur, proveedor_id, "combustible")
                    for row_number, row in rows:
                        if not any(str(value or "").strip() for value in row.values()):
                            skipped += 1
                            continue
                        vehiculo_hint = str(row.get("vehiculo") or row.get("codigo") or row.get("codigo vehiculo") or row.get("movil") or row.get("nro movil") or "").strip()
                        fecha_hint = str(row.get("fecha real") or row.get("fecha") or row.get("fecha carga") or "").strip()
                        producto_hint = str(row.get("producto") or row.get("tipo combustible") or row.get("combustible") or "").strip()
                        if not vehiculo_hint and not fecha_hint:
                            skipped += 1
                            continue
                        if not _is_combustible_import_product(producto_hint):
                            skipped += 1
                            continue
                        try:
                            mapped = self._map_combustible_import_row(row)
                            normalized_factura = _normalize_invoice_number(mapped["nro_factura"])
                            if normalized_factura:
                                factura_key = f"{proveedor_id or 0}:{normalized_factura}"
                                first_row = seen_facturas.get(factura_key)
                                if first_row:
                                    raise ValueError(f"La factura {mapped['nro_factura']} esta repetida en el archivo (fila {first_row}).")
                                seen_facturas[factura_key] = row_number
                            vehiculo = self._find_vehiculo_flota_import(cur, mapped["vehiculo_ref"])
                            self._validate_flota_sucursal_scope(vehiculo, sucursal_scope)
                            self._validate_factura_combustible_unica(cur, proveedor_id, mapped["nro_factura"])
                            semana, anho = _iso_week_parts(mapped["fecha"])
                            precio_litro = round(mapped["importe"] / mapped["litros"], 2)
                            cur.execute(
                                """
                                INSERT INTO cargas_combustible(
                                    vehiculo_id, fecha, proveedor_id, litros, importe, precio_litro,
                                    tipo_combustible, km_actual, nro_factura, observacion, semana, anho, cargado_por
                                )
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                """,
                                (
                                    vehiculo["id"],
                                    mapped["fecha"],
                                    proveedor_id,
                                    mapped["litros"],
                                    mapped["importe"],
                                    precio_litro,
                                    mapped["tipo_combustible"] or None,
                                    None,
                                    mapped["nro_factura"] or None,
                                    mapped["observacion"] or None,
                                    semana,
                                    anho,
                                    str(cargado_por or "").strip() or None,
                                ),
                            )
                            inserted += 1
                        except Exception as exc:
                            error_text = str(exc)
                            if error_text.startswith("No existe un vehiculo que coincida con '") and error_text.endswith("'."):
                                vehiculo_ref = error_text[len("No existe un vehiculo que coincida con '") : -2]
                                missing_vehicles[vehiculo_ref] = missing_vehicles.get(vehiculo_ref, 0) + 1
                            errors.append({"row": row_number, "error": str(exc)})
                conn.commit()
            except Exception:
                conn.rollback()
                raise

        return {
            "inserted": inserted,
            "skipped": skipped,
            "errors": errors[:50],
            "missing_vehicles": [
                {"vehiculo": vehiculo, "count": count}
                for vehiculo, count in sorted(missing_vehicles.items(), key=lambda item: (-item[1], item[0].lower()))
            ],
        }

    def preview_cargas_combustible_import(self, payload, sucursal_scope=None):
        file_name = str(payload.get("file_name") or "").strip()
        file_content = str(payload.get("file_content") or "").strip()
        proveedor_id = payload.get("proveedor_id")
        proveedor_id = int(proveedor_id) if str(proveedor_id or "").strip() else None
        if not file_name or not file_content:
            raise ValueError("Debes adjuntar un archivo CSV o XLSX.")

        rows = self._read_combustible_import_rows(file_name, file_content)
        if not rows:
            raise ValueError("No se encontraron filas para previsualizar.")

        items = []
        skipped = 0
        ok_count = 0
        error_count = 0
        seen_facturas: dict[str, int] = {}
        with self._connect(readonly=True) as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                self._validate_proveedor_tipo(cur, proveedor_id, "combustible")
                for row_number, row in rows:
                    if not any(str(value or "").strip() for value in row.values()):
                        skipped += 1
                        continue
                    vehiculo_hint = str(row.get("vehiculo") or row.get("codigo") or row.get("codigo vehiculo") or row.get("movil") or row.get("nro movil") or "").strip()
                    fecha_hint = str(row.get("fecha real") or row.get("fecha") or row.get("fecha carga") or "").strip()
                    producto_hint = str(row.get("producto") or row.get("tipo combustible") or row.get("combustible") or "").strip()
                    if not vehiculo_hint and not fecha_hint:
                        skipped += 1
                        continue
                    if not _is_combustible_import_product(producto_hint):
                        skipped += 1
                        continue
                    try:
                        mapped = self._map_combustible_import_row(row)
                        normalized_factura = _normalize_invoice_number(mapped["nro_factura"])
                        if normalized_factura:
                            factura_key = f"{proveedor_id or 0}:{normalized_factura}"
                            first_row = seen_facturas.get(factura_key)
                            if first_row:
                                raise ValueError(f"La factura {mapped['nro_factura']} esta repetida en el archivo (fila {first_row}).")
                            seen_facturas[factura_key] = row_number
                        vehiculo = self._find_vehiculo_flota_import(cur, mapped["vehiculo_ref"])
                        self._validate_flota_sucursal_scope(vehiculo, sucursal_scope)
                        self._validate_factura_combustible_unica(cur, proveedor_id, mapped["nro_factura"])
                        precio_litro = round(mapped["importe"] / mapped["litros"], 2)
                        items.append(
                            {
                                "row": row_number,
                                "vehiculo_ref": mapped["vehiculo_ref"],
                                "vehiculo_match": self._vehiculo_flota_label(vehiculo),
                                "fecha": mapped["fecha"],
                                "tipo_combustible": mapped["tipo_combustible"],
                                "litros": mapped["litros"],
                                "precio_litro": precio_litro,
                                "importe": mapped["importe"],
                                "nro_factura": mapped["nro_factura"],
                                "status": "ok",
                            }
                        )
                        ok_count += 1
                    except Exception as exc:
                        items.append(
                            {
                                "row": row_number,
                                "vehiculo_ref": vehiculo_hint,
                                "vehiculo_match": "",
                                "fecha": fecha_hint,
                                "tipo_combustible": str(row.get("producto") or row.get("tipo combustible") or row.get("combustible") or ""),
                                "litros": 0,
                                "precio_litro": 0,
                                "importe": 0,
                                "nro_factura": str(row.get("codigo de autorizacion") or row.get("autorizacion") or row.get("factura") or row.get("nro factura") or ""),
                                "status": "error",
                                "error": str(exc),
                            }
                        )
                        error_count += 1
        return {
            "items": items[:120],
            "skipped": skipped,
            "ok_count": ok_count,
            "error_count": error_count,
        }

    def list_gastos_flota(self, desde=None, hasta=None, vehiculo_id=None, tipo_gasto_id=None, sucursal=None):
        filters = []
        params: list[Any] = []
        if desde:
            filters.append("g.fecha >= %s")
            params.append(desde)
        if hasta:
            filters.append("g.fecha <= %s")
            params.append(hasta)
        if vehiculo_id:
            filters.append("g.vehiculo_id = %s")
            params.append(int(vehiculo_id))
        if tipo_gasto_id:
            filters.append("g.tipo_gasto_id = %s")
            params.append(int(tipo_gasto_id))
        if sucursal:
            filters.append("LOWER(COALESCE(v.sucursal, '')) = %s")
            params.append(str(sucursal).strip().lower())
        where = f"WHERE {' AND '.join(filters)}" if filters else ""

        with self._connect(readonly=True) as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    f"""
                    SELECT g.id,
                           g.fecha,
                           g.vehiculo_id,
                           v.codigo AS vehiculo_codigo,
                           v.nombre AS vehiculo_nombre,
                           v.chapa,
                           v.sucursal,
                           g.tipo_gasto_id,
                           tg.nombre AS tipo_gasto,
                           tg.requiere_km,
                           g.proveedor_id,
                           COALESCE(NULLIF(g.proveedor_nombre, ''), p.nombre, '') AS proveedor_nombre,
                           COALESCE(g.proveedor_ruc, '') AS proveedor_ruc,
                           g.importe,
                           g.km_actual,
                           COALESCE(g.nro_factura, '') AS nro_factura,
                           COALESCE(g.detalle, '') AS detalle,
                           g.semana,
                           g.anho,
                           g.cargado_por,
                           g.creado_en,
                           g.eliminado_en,
                           COALESCE(g.eliminado_por, '') AS eliminado_por,
                           COALESCE(g.motivo_eliminacion, '') AS motivo_eliminacion
                    FROM gastos_flota g
                    JOIN vehiculos v ON v.id = g.vehiculo_id
                    JOIN tipos_gasto_flota tg ON tg.id = g.tipo_gasto_id
                    LEFT JOIN proveedores_flota p ON p.id = g.proveedor_id
                    {where}
                    ORDER BY g.fecha DESC, g.id DESC
                    LIMIT 250
                    """,
                    params,
                )
                rows = cur.fetchall()
        return {"items": rows}

    def save_gasto_flota(self, payload, cargado_por=None, sucursal_scope=None, user_role=None):
        gasto_id = payload.get("id")
        gasto_id = int(gasto_id) if str(gasto_id or "").strip() else None
        fecha = _parse_date(payload.get("fecha"))
        vehiculo_id = int(payload.get("vehiculo_id") or 0)
        tipo_gasto_id = int(payload.get("tipo_gasto_id") or 0)
        proveedor_id = payload.get("proveedor_id")
        proveedor_id = int(proveedor_id) if str(proveedor_id or "").strip() else None
        proveedor_nombre = str(payload.get("proveedor_nombre") or "").strip()
        proveedor_ruc = str(payload.get("proveedor_ruc") or "").strip()
        importe = _parse_number(payload.get("importe") or 0)
        km_actual_raw = payload.get("km_actual")
        km_actual = _parse_number(km_actual_raw) if str(km_actual_raw or "").strip() else None
        nro_factura = str(payload.get("nro_factura") or "").strip()
        detalle = str(payload.get("detalle") or "").strip()

        if not fecha:
            raise ValueError("La fecha es obligatoria.")
        if vehiculo_id <= 0:
            raise ValueError("Debes seleccionar un vehiculo.")
        if tipo_gasto_id <= 0:
            raise ValueError("Debes seleccionar un tipo de gasto.")
        if importe <= 0:
            raise ValueError("El importe debe ser mayor a 0.")

        semana, anho = _iso_week_parts(fecha)

        with self._connect(readonly=False) as conn:
            try:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    vehiculo = self._require_vehiculo_activo(cur, vehiculo_id)
                    self._validate_flota_sucursal_scope(vehiculo, sucursal_scope)
                    if gasto_id:
                        self._validate_gasto_flota_editable(cur, gasto_id, user_role, sucursal_scope)
                    tipo = self._get_tipo_gasto_flota(cur, tipo_gasto_id)
                    self._validate_km_flota(cur, vehiculo_id, km_actual)
                    if proveedor_id:
                        self._validate_proveedor_tipo(cur, proveedor_id, None)
                    if not proveedor_id and not proveedor_nombre:
                        raise ValueError("Ingresa el nombre del proveedor o selecciona uno existente.")
                    if proveedor_id and not proveedor_nombre:
                        cur.execute("SELECT nombre, COALESCE(ruc, '') AS ruc FROM proveedores_flota WHERE id = %s", (proveedor_id,))
                        proveedor_row = cur.fetchone() or {}
                        proveedor_nombre = str(proveedor_row.get("nombre") or "")
                        if not proveedor_ruc:
                            proveedor_ruc = str(proveedor_row.get("ruc") or "")
                    self._validate_factura_gasto_unica(cur, proveedor_id, proveedor_nombre, proveedor_ruc, nro_factura, exclude_id=gasto_id)
                    if gasto_id:
                        cur.execute(
                            """
                            UPDATE gastos_flota
                            SET vehiculo_id = %s,
                                fecha = %s,
                                tipo_gasto_id = %s,
                                proveedor_id = %s,
                                importe = %s,
                                km_actual = %s,
                                nro_factura = %s,
                                detalle = %s,
                                semana = %s,
                                anho = %s,
                                proveedor_nombre = %s,
                                proveedor_ruc = %s
                            WHERE id = %s
                              AND eliminado_en IS NULL
                            RETURNING id, fecha, vehiculo_id, tipo_gasto_id, proveedor_id, importe, km_actual,
                                      nro_factura, detalle, semana, anho, cargado_por, creado_en, proveedor_nombre, proveedor_ruc
                            """,
                            (
                                vehiculo_id,
                                fecha,
                                tipo_gasto_id,
                                proveedor_id,
                                importe,
                                km_actual,
                                nro_factura,
                                detalle,
                                semana,
                                anho,
                                proveedor_nombre or None,
                                proveedor_ruc or None,
                                gasto_id,
                            ),
                        )
                    else:
                        cur.execute(
                            """
                            INSERT INTO gastos_flota(
                                vehiculo_id, fecha, tipo_gasto_id, proveedor_id, importe, km_actual,
                                nro_factura, detalle, semana, anho, cargado_por, proveedor_nombre, proveedor_ruc
                            )
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            RETURNING id, fecha, vehiculo_id, tipo_gasto_id, proveedor_id, importe, km_actual,
                                      nro_factura, detalle, semana, anho, cargado_por, creado_en, proveedor_nombre, proveedor_ruc
                            """,
                            (
                                vehiculo_id,
                                fecha,
                                tipo_gasto_id,
                                proveedor_id,
                                importe,
                                km_actual,
                                nro_factura,
                                detalle,
                                semana,
                                anho,
                                str(cargado_por or "").strip() or None,
                                proveedor_nombre or None,
                                proveedor_ruc or None,
                            ),
                        )
                    row = cur.fetchone()
                    if not row:
                        raise ValueError("Gasto no encontrado.")
                    row["vehiculo_nombre"] = vehiculo["nombre"]
                    row["vehiculo_codigo"] = vehiculo["codigo"]
                    row["chapa"] = vehiculo["chapa"]
                    row["sucursal"] = vehiculo["sucursal"]
                    row["tipo_gasto"] = tipo["nombre"]
                    row["proveedor_nombre"] = proveedor_nombre
                    row["proveedor_ruc"] = proveedor_ruc
                conn.commit()
                return row
            except psycopg2.errors.UniqueViolation:
                conn.rollback()
                raise ValueError("La factura ya esta registrada en gastos de flota.")
            except Exception:
                conn.rollback()
                raise

    def _validate_gasto_flota_editable(self, cur, gasto_id, user_role=None, sucursal_scope=None):
        cur.execute(
            """
            SELECT g.id,
                   g.creado_en,
                   g.eliminado_en,
                   v.sucursal
            FROM gastos_flota g
            JOIN vehiculos v ON v.id = g.vehiculo_id
            WHERE g.id = %s
            """,
            (int(gasto_id),),
        )
        row = cur.fetchone()
        if not row:
            raise ValueError("El gasto no existe.")
        self._validate_flota_sucursal_scope(row, sucursal_scope)
        if row.get("eliminado_en"):
            raise ValueError("El gasto ya fue eliminado.")
        if str(user_role or "").strip().lower() == "recepcion":
            cur.execute("SELECT %s::timestamp >= NOW() - INTERVAL '2 days' AS editable", (row["creado_en"],))
            editable = bool((cur.fetchone() or {}).get("editable"))
            if not editable:
                raise PermissionDenied("Recepcion solo puede editar o eliminar gastos creados en los ultimos 2 dias.")
        return row

    def delete_gasto_flota(self, payload, eliminado_por=None, sucursal_scope=None, user_role=None):
        gasto_id = _parse_int(payload.get("id") or 0)
        motivo = str(payload.get("motivo") or "").strip()

        if gasto_id <= 0:
            raise ValueError("Falta el id del gasto.")
        if not motivo:
            raise ValueError("Debes indicar el motivo de eliminacion.")

        with self._connect(readonly=False) as conn:
            try:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    self._validate_gasto_flota_editable(cur, gasto_id, user_role, sucursal_scope)
                    cur.execute(
                        """
                        UPDATE gastos_flota
                        SET eliminado_en = NOW(),
                            eliminado_por = %s,
                            motivo_eliminacion = %s
                        WHERE id = %s
                        RETURNING id, eliminado_en, COALESCE(eliminado_por, '') AS eliminado_por, COALESCE(motivo_eliminacion, '') AS motivo_eliminacion
                        """,
                        (str(eliminado_por or "").strip() or None, motivo, gasto_id),
                    )
                    deleted = cur.fetchone()
                conn.commit()
                return {"ok": True, "item": deleted}
            except Exception:
                conn.rollback()
                raise

    def get_flota_resumen_semanal(self, semana=None, anho=None, vehiculo_id=None, sucursal=None):
        today = date.today()
        mes = int(semana or today.month)
        anho = int(anho or today.year)
        filters = ["w.mes = %s", "w.anho = %s"]
        params: list[Any] = [mes, anho]
        if vehiculo_id:
            filters.append("v.id = %s")
            params.append(int(vehiculo_id))
        if sucursal:
            filters.append("LOWER(COALESCE(v.sucursal, '')) = %s")
            params.append(str(sucursal).strip().lower())
        where = f"WHERE {' AND '.join(filters)}"

        with self._connect(readonly=True) as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    f"""
                    WITH combustible AS (
                        SELECT vehiculo_id,
                               EXTRACT(MONTH FROM fecha)::int AS mes,
                               anho,
                               COALESCE(SUM(litros), 0)::numeric AS litros,
                               COALESCE(SUM(importe), 0)::numeric AS combustible_total,
                               MAX(km_actual) AS km_max,
                               MIN(km_actual) AS km_min
                        FROM cargas_combustible
                        WHERE eliminado_en IS NULL
                        GROUP BY vehiculo_id, EXTRACT(MONTH FROM fecha), anho
                    ),
                    gastos AS (
                        SELECT vehiculo_id,
                               EXTRACT(MONTH FROM fecha)::int AS mes,
                               anho,
                               COALESCE(SUM(importe), 0)::numeric AS otros_gastos
                        FROM gastos_flota
                        WHERE eliminado_en IS NULL
                        GROUP BY vehiculo_id, EXTRACT(MONTH FROM fecha), anho
                    ),
                    week_data AS (
                        SELECT COALESCE(c.vehiculo_id, g.vehiculo_id) AS vehiculo_id,
                               COALESCE(c.mes, g.mes) AS mes,
                               COALESCE(c.anho, g.anho) AS anho,
                               COALESCE(c.litros, 0)::numeric AS litros,
                               COALESCE(c.combustible_total, 0)::numeric AS combustible_total,
                               COALESCE(g.otros_gastos, 0)::numeric AS otros_gastos,
                               c.km_max,
                               c.km_min
                        FROM combustible c
                        FULL OUTER JOIN gastos g
                          ON g.vehiculo_id = c.vehiculo_id
                         AND g.mes = c.mes
                         AND g.anho = c.anho
                    )
                    SELECT v.id AS vehiculo_id,
                           v.codigo,
                           v.nombre,
                           v.chapa,
                           v.sucursal,
                           v.tipo,
                           w.mes,
                           w.anho,
                           w.litros,
                           w.combustible_total,
                           w.otros_gastos,
                           (w.combustible_total + w.otros_gastos)::numeric AS total_general,
                           CASE
                               WHEN COALESCE(w.litros, 0) > 0 THEN ROUND((w.combustible_total / w.litros)::numeric, 2)
                               ELSE 0
                           END AS precio_litro_promedio,
                           w.km_min,
                           w.km_max AS km_actual,
                           CASE
                               WHEN w.km_max IS NOT NULL AND w.km_min IS NOT NULL AND w.km_max >= w.km_min
                                   THEN ROUND((w.km_max - w.km_min)::numeric, 2)
                               ELSE 0
                           END AS km_recorrido,
                           CASE
                               WHEN w.km_max IS NOT NULL AND w.km_min IS NOT NULL AND w.km_max > w.km_min
                                   THEN ROUND(((w.combustible_total + w.otros_gastos) / NULLIF((w.km_max - w.km_min), 0))::numeric, 2)
                               ELSE NULL
                           END AS costo_por_km
                    FROM week_data w
                    JOIN vehiculos v ON v.id = w.vehiculo_id
                    {where}
                    ORDER BY total_general DESC, v.nombre
                    """,
                    params,
                )
                items = cur.fetchall()

                comparison_points = []
                for offset in range(-3, 1):
                    comp_month, comp_year = _shift_month(mes, anho, offset)
                    cur.execute(
                        """
                        WITH combustible AS (
                            SELECT vehiculo_id, EXTRACT(MONTH FROM fecha)::int AS mes, anho, COALESCE(SUM(importe), 0)::numeric AS combustible_total
                            FROM cargas_combustible
                            WHERE eliminado_en IS NULL
                            GROUP BY vehiculo_id, EXTRACT(MONTH FROM fecha), anho
                        ),
                        gastos AS (
                            SELECT vehiculo_id, EXTRACT(MONTH FROM fecha)::int AS mes, anho, COALESCE(SUM(importe), 0)::numeric AS otros_gastos
                            FROM gastos_flota
                            WHERE eliminado_en IS NULL
                            GROUP BY vehiculo_id, EXTRACT(MONTH FROM fecha), anho
                        ),
                        week_data AS (
                            SELECT COALESCE(c.vehiculo_id, g.vehiculo_id) AS vehiculo_id,
                                   COALESCE(c.mes, g.mes) AS mes,
                                   COALESCE(c.anho, g.anho) AS anho,
                                   COALESCE(c.combustible_total, 0)::numeric AS combustible_total,
                                   COALESCE(g.otros_gastos, 0)::numeric AS otros_gastos
                            FROM combustible c
                            FULL OUTER JOIN gastos g
                              ON g.vehiculo_id = c.vehiculo_id
                             AND g.mes = c.mes
                             AND g.anho = c.anho
                        )
                        SELECT %s AS mes,
                               %s AS anho,
                               COUNT(*)::int AS vehiculos,
                               COALESCE(SUM(w.combustible_total), 0)::numeric AS combustible_total,
                               COALESCE(SUM(w.otros_gastos), 0)::numeric AS otros_gastos,
                               COALESCE(SUM(w.combustible_total + w.otros_gastos), 0)::numeric AS total_general
                        FROM week_data w
                        JOIN vehiculos v ON v.id = w.vehiculo_id
                        WHERE w.mes = %s
                          AND w.anho = %s
                          AND (%s::int IS NULL OR v.id = %s::int)
                          AND (%s::text IS NULL OR LOWER(COALESCE(v.sucursal, '')) = %s::text)
                        """,
                        (
                            comp_month,
                            comp_year,
                            comp_month,
                            comp_year,
                            int(vehiculo_id) if vehiculo_id else None,
                            int(vehiculo_id) if vehiculo_id else None,
                            str(sucursal).strip().lower() if sucursal else None,
                            str(sucursal).strip().lower() if sucursal else None,
                        ),
                    )
                    comparison_points.append(cur.fetchone() or {})

        totales = {
            "mes": mes,
            "anho": anho,
            "vehiculos": len(items),
            "litros": sum(float(item.get("litros") or 0) for item in items),
            "combustible_total": sum(float(item.get("combustible_total") or 0) for item in items),
            "otros_gastos": sum(float(item.get("otros_gastos") or 0) for item in items),
            "total_general": sum(float(item.get("total_general") or 0) for item in items),
        }
        sucursales = []
        ranking = []
        for item in items:
            slug = str(item.get("sucursal") or "").strip().lower()
            found = next((row for row in sucursales if row["sucursal"] == slug), None)
            if not found:
                found = {
                    "sucursal": slug,
                    "vehiculos": 0,
                    "litros": 0.0,
                    "combustible_total": 0.0,
                    "otros_gastos": 0.0,
                    "total_general": 0.0,
                }
                sucursales.append(found)
            found["vehiculos"] += 1
            found["litros"] += float(item.get("litros") or 0)
            found["combustible_total"] += float(item.get("combustible_total") or 0)
            found["otros_gastos"] += float(item.get("otros_gastos") or 0)
            found["total_general"] += float(item.get("total_general") or 0)

        ranking = sorted(
            [
                {
                    "vehiculo_id": item["vehiculo_id"],
                    "codigo": item["codigo"],
                    "nombre": item["nombre"],
                    "sucursal": item.get("sucursal"),
                    "tipo": item.get("tipo"),
                    "total_general": item.get("total_general"),
                    "costo_por_km": item.get("costo_por_km"),
                }
                for item in items
            ],
            key=lambda row: float(row.get("total_general") or 0),
            reverse=True,
        )[:5]

        return {
            "totales": totales,
            "items": items,
            "totalesPorSucursal": sorted(sucursales, key=lambda row: float(row["total_general"]), reverse=True),
            "comparativoMeses": comparison_points,
            "rankingCosto": ranking,
        }

    def list_cargas_combustible_eliminadas(self, mes=None, anho=None, vehiculo_id=None, sucursal=None):
        filters = ["EXTRACT(MONTH FROM c.fecha)::int = %s", "c.anho = %s", "c.eliminado_en IS NOT NULL"]
        params: list[Any] = [int(mes), int(anho)]
        if vehiculo_id:
            filters.append("c.vehiculo_id = %s")
            params.append(int(vehiculo_id))
        if sucursal:
            filters.append("LOWER(COALESCE(v.sucursal, '')) = %s")
            params.append(str(sucursal).strip().lower())
        where = f"WHERE {' AND '.join(filters)}"

        with self._connect(readonly=True) as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    f"""
                    SELECT c.fecha,
                           c.litros,
                           c.importe,
                           c.nro_factura,
                           COALESCE(c.motivo_eliminacion, '') AS motivo_eliminacion,
                           COALESCE(c.eliminado_por, '') AS eliminado_por,
                           c.eliminado_en,
                           v.codigo,
                           v.nombre,
                           v.chapa,
                           v.sucursal
                    FROM cargas_combustible c
                    JOIN vehiculos v ON v.id = c.vehiculo_id
                    {where}
                    ORDER BY c.eliminado_en DESC, c.id DESC
                    """,
                    params,
                )
                return cur.fetchall()

    def build_flota_resumen_mensual_pdf(self, mes=None, anho=None, vehiculo_id=None, sucursal=None):
        if SimpleDocTemplate is None:
            raise RuntimeError("ReportLab no esta instalado. Instale reportlab para generar PDF.")

        resumen = self.get_flota_resumen_semanal(semana=mes, anho=anho, vehiculo_id=vehiculo_id, sucursal=sucursal)
        totales = resumen.get("totales") or {}
        items = resumen.get("items") or []
        totales_por_sucursal = resumen.get("totalesPorSucursal") or []
        ranking = resumen.get("rankingCosto") or []
        report_month = int(totales.get("mes") or mes or date.today().month)
        report_year = int(totales.get("anho") or anho or date.today().year)
        eliminadas = self.list_cargas_combustible_eliminadas(report_month, report_year, vehiculo_id=vehiculo_id, sucursal=sucursal)

        month_name = datetime(report_year, report_month, 1).strftime("%B %Y").capitalize()
        filters_applied = []
        if sucursal:
            sucursal_meta = SUCURSALES.get(str(sucursal).strip().lower())
            filters_applied.append(f"Sucursal: {sucursal_meta['nombre'] if sucursal_meta else str(sucursal)}")
        if vehiculo_id:
            filters_applied.append(f"Vehiculo ID: {int(vehiculo_id)}")
        filter_text = " | ".join(filters_applied) if filters_applied else "Sin filtros adicionales"

        def fmt_gs(value, dec=0):
            return f"Gs. {_fmt_float(value, dec)}"

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "FlotaMonthlyTypeTitle",
            parent=styles["Title"],
            fontName="Courier-Bold",
            fontSize=15,
            leading=17,
            spaceAfter=4,
        )
        heading_style = ParagraphStyle(
            "FlotaMonthlyTypeHeading",
            parent=styles["Heading3"],
            fontName="Courier-Bold",
            fontSize=9,
            leading=10,
            spaceBefore=2,
            spaceAfter=4,
        )
        body_style = ParagraphStyle(
            "FlotaMonthlyTypeBody",
            parent=styles["Normal"],
            fontName="Courier",
            fontSize=8,
            leading=9,
            spaceBefore=0,
            spaceAfter=2,
        )
        subtotal_style = ParagraphStyle(
            "FlotaMonthlyTypeSubtotal",
            parent=styles["Normal"],
            fontName="Courier-Bold",
            fontSize=7,
            leading=8,
            alignment=1,
            spaceBefore=0,
            spaceAfter=0,
        )
        subtotal_label_style = ParagraphStyle(
            "FlotaMonthlyTypeSubtotalLabel",
            parent=subtotal_style,
            alignment=0,
            wordWrap="CJK",
        )
        kpi_value_style = ParagraphStyle(
            "FlotaMonthlyTypeKpiValue",
            parent=styles["Normal"],
            fontName="Courier-Bold",
            fontSize=13,
            leading=15,
            alignment=1,
            spaceBefore=0,
            spaceAfter=0,
        )

        kpi_data = [[
            self._pdf_wrap_cell_typewriter("Vehiculos", align="CENTER", font_size=8, leading=9),
            self._pdf_wrap_cell_typewriter("Combustible", align="CENTER", font_size=8, leading=9),
            self._pdf_wrap_cell_typewriter("Gastos", align="CENTER", font_size=8, leading=9),
            self._pdf_wrap_cell_typewriter("Monto total", align="CENTER", font_size=8, leading=9),
        ], [
            Paragraph(_fmt_int(totales.get("vehiculos", 0)), kpi_value_style),
            Paragraph(fmt_gs(totales.get("combustible_total", 0)), kpi_value_style),
            Paragraph(fmt_gs(totales.get("otros_gastos", 0)), kpi_value_style),
            Paragraph(fmt_gs(totales.get("total_general", 0)), kpi_value_style),
        ]]

        story = [
            Paragraph("<b>Reporte mensual de flota</b>", title_style),
            Paragraph(
                f"Periodo: <b>{escape(month_name)}</b> | Generado: <b>{datetime.now().strftime('%Y-%m-%d %H:%M')}</b>",
                body_style,
            ),
            Paragraph(filter_text, body_style),
            Spacer(0, 6),
            self._build_table_compact_typewriter(kpi_data, col_widths=[120, 150, 150, 150], header_fill=colors.HexColor("#FFF3CD")),
            Spacer(0, 8),
        ]

        data_sucursal = [["Sucursal", "Vehiculos", "Monto combustible", "Gastos", "Monto total"]]
        for row in totales_por_sucursal:
            sucursal_slug = str(row.get("sucursal") or "").strip().lower()
            sucursal_meta = SUCURSALES.get(sucursal_slug)
            data_sucursal.append([
                sucursal_meta["nombre"] if sucursal_meta else (sucursal_slug or "Sin sucursal"),
                _fmt_int(row.get("vehiculos", 0)),
                fmt_gs(row.get("combustible_total", 0)),
                fmt_gs(row.get("otros_gastos", 0)),
                fmt_gs(row.get("total_general", 0)),
            ])
        if len(data_sucursal) == 1:
            data_sucursal.append(["Sin datos", "0", "0", "0", "0"])

        story.append(Paragraph("<b>Resumen por sucursal</b>", heading_style))
        story.append(self._build_table_compact_typewriter(data_sucursal, col_widths=[120, 80, 120, 120, 120]))
        story.append(Spacer(0, 8))

        def sucursal_label(value):
            slug = str(value or "").strip().lower()
            meta = SUCURSALES.get(slug)
            return meta["nombre"] if meta else (slug or "Sin sucursal")

        def tipo_label(value):
            return str(value or "").strip() or "Sin tipo"

        resumen_tipo = []
        for row in items:
            sucursal_name = sucursal_label(row.get("sucursal"))
            tipo_name = tipo_label(row.get("tipo"))
            found = next((item for item in resumen_tipo if item["tipo"] == tipo_name), None)
            if not found:
                found = {
                    "tipo": tipo_name,
                    "sucursales": set(),
                    "vehiculos": 0,
                    "litros": 0.0,
                    "combustible_total": 0.0,
                    "otros_gastos": 0.0,
                    "total_general": 0.0,
                }
                resumen_tipo.append(found)
            found["sucursales"].add(sucursal_name)
            found["vehiculos"] += 1
            found["litros"] += float(row.get("litros") or 0)
            found["combustible_total"] += float(row.get("combustible_total") or 0)
            found["otros_gastos"] += float(row.get("otros_gastos") or 0)
            found["total_general"] += float(row.get("total_general") or 0)

        data_tipo = [["Tipo de vehiculo", "Sucursales", "Vehiculos", "Combustible", "Gastos", "Monto total", "Litros"]]
        for row in sorted(resumen_tipo, key=lambda item: item["tipo"]):
            data_tipo.append([
                self._pdf_wrap_cell_typewriter(row["tipo"], align="LEFT", font_size=7, leading=8),
                _fmt_int(len(row["sucursales"])),
                _fmt_int(row["vehiculos"]),
                fmt_gs(row["combustible_total"]),
                fmt_gs(row["otros_gastos"]),
                fmt_gs(row["total_general"]),
                _fmt_float(row["litros"], 2),
            ])
        if len(data_tipo) == 1:
            data_tipo.append(["Sin datos", "0", "0", "0", "0", "0", "0,00"])

        story.append(Paragraph("<b>Resumen por tipo de vehiculo</b>", heading_style))
        story.append(self._build_table_compact_typewriter(data_tipo, col_widths=[185, 65, 65, 95, 75, 85, 60]))

        story.append(Spacer(0, 8))
        story.append(Paragraph("<b>Detalle por tipo de vehiculo</b>", heading_style))
        grouped_items: dict[tuple[str, str], list[dict[str, Any]]] = {}
        for row in items:
            key = (tipo_label(row.get("tipo")), sucursal_label(row.get("sucursal")))
            grouped_items.setdefault(key, []).append(row)

        if not grouped_items:
            story.append(self._build_table_compact_typewriter([["Tipo", "Sucursal", "Vehiculo", "Combustible", "Gastos", "Monto total", "Litros"], ["Sin datos", "-", "-", "0", "0", "0", "0,00"]], col_widths=[120, 80, 180, 90, 70, 75, 55]))
        for (tipo_name, sucursal_name), rows in sorted(grouped_items.items(), key=lambda item: (item[0][0], item[0][1])):
            subtotal_combustible = sum(float(row.get("combustible_total") or 0) for row in rows)
            subtotal_gastos = sum(float(row.get("otros_gastos") or 0) for row in rows)
            subtotal_total = sum(float(row.get("total_general") or 0) for row in rows)
            subtotal_litros = sum(float(row.get("litros") or 0) for row in rows)
            story.append(Spacer(0, 5))
            story.append(Paragraph(f"<b>{escape(tipo_name)} - {escape(sucursal_name)}</b>", body_style))
            data_vehiculos = [["Vehiculo", "Monto combustible", "Gastos", "Monto total", "Litros"]]
            for row in sorted(rows, key=lambda item: float(item.get("total_general") or 0), reverse=True):
                data_vehiculos.append([
                    self._pdf_wrap_cell_typewriter(self._vehiculo_flota_label(row), align="LEFT", font_size=7, leading=8),
                    fmt_gs(row.get("combustible_total", 0)),
                    fmt_gs(row.get("otros_gastos", 0)),
                    fmt_gs(row.get("total_general", 0)),
                    _fmt_float(row.get("litros", 0), 2),
                ])
            data_vehiculos.append([
                Paragraph("Subtotal", subtotal_label_style),
                Paragraph(fmt_gs(subtotal_combustible), subtotal_style),
                Paragraph(fmt_gs(subtotal_gastos), subtotal_style),
                Paragraph(fmt_gs(subtotal_total), subtotal_style),
                Paragraph(_fmt_float(subtotal_litros, 2), subtotal_style),
            ])
            story.append(self._build_table_compact_typewriter(data_vehiculos, col_widths=[300, 110, 80, 85, 65]))

        if ranking:
            story.append(Spacer(0, 8))
            data_ranking = [["Vehiculo", "Sucursal", "Monto total", "Costo por km"]]
            for row in ranking:
                sucursal_slug = str(row.get("sucursal") or "").strip().lower()
                sucursal_meta = SUCURSALES.get(sucursal_slug)
                costo_por_km = row.get("costo_por_km")
                data_ranking.append([
                    self._pdf_wrap_cell_typewriter(self._vehiculo_flota_label(row), align="LEFT", font_size=7, leading=8),
                    sucursal_meta["nombre"] if sucursal_meta else (sucursal_slug or "Sin sucursal"),
                    fmt_gs(row.get("total_general", 0)),
                    fmt_gs(costo_por_km, 2) if costo_por_km is not None else "-",
                ])
            story.append(Paragraph("<b>Ranking de costo</b>", heading_style))
            story.append(self._build_table_compact_typewriter(data_ranking, col_widths=[250, 100, 100, 90]))

        if eliminadas:
            story.append(Spacer(0, 8))
            data_eliminadas = [["Fecha", "Vehiculo", "Litros", "Importe", "Factura", "Motivo", "Eliminado por", "Eliminado en"]]
            for row in eliminadas:
                eliminado_en = row.get("eliminado_en")
                eliminado_en_text = eliminado_en.strftime("%Y-%m-%d %H:%M") if isinstance(eliminado_en, datetime) else str(eliminado_en or "")
                data_eliminadas.append([
                    str(row.get("fecha") or ""),
                    self._pdf_wrap_cell_typewriter(self._vehiculo_flota_label(row), align="LEFT", font_size=7, leading=8),
                    _fmt_float(row.get("litros", 0), 2),
                    fmt_gs(row.get("importe", 0)),
                    str(row.get("nro_factura") or "-"),
                    self._pdf_wrap_cell_typewriter(str(row.get("motivo_eliminacion") or "-"), align="LEFT", font_size=7, leading=8),
                    str(row.get("eliminado_por") or "-"),
                    eliminado_en_text or "-",
                ])
            story.append(Paragraph("<b>Cargas de combustible eliminadas</b>", heading_style))
            story.append(self._build_table_compact_typewriter(data_eliminadas, col_widths=[60, 150, 55, 70, 70, 160, 80, 80]))

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=18,
            rightMargin=18,
            topMargin=18,
            bottomMargin=18,
        )
        doc.build(story)
        return buffer.getvalue(), f"Reporte_Flota_{report_year}_{report_month:02d}.pdf"

    def _require_vehiculo_activo(self, cur, vehiculo_id):
        cur.execute(
            """
            SELECT id, codigo, chapa, nombre, sucursal, activo
            FROM vehiculos
            WHERE id = %s
            """,
            (int(vehiculo_id),),
        )
        row = cur.fetchone()
        if not row:
            raise ValueError("Vehiculo no encontrado.")
        if not bool(row["activo"]):
            raise ValueError("El vehiculo esta inactivo y no admite nuevos movimientos.")
        return row

    def _validate_flota_sucursal_scope(self, vehiculo, sucursal_scope):
        scoped_sucursal = str(sucursal_scope or "").strip().lower()
        if not scoped_sucursal:
            return
        vehiculo_sucursal = str((vehiculo or {}).get("sucursal") or "").strip().lower()
        if vehiculo_sucursal != scoped_sucursal:
            raise PermissionDenied("No tienes permisos para operar vehiculos de otra sucursal.")

    def _validate_proveedor_tipo(self, cur, proveedor_id, expected_type):
        if proveedor_id is None:
            return
        cur.execute(
            """
            SELECT id, tipo, activo
            FROM proveedores_flota
            WHERE id = %s
            """,
            (int(proveedor_id),),
        )
        row = cur.fetchone()
        if not row:
            raise ValueError("Proveedor no encontrado.")
        if not bool(row["activo"]):
            raise ValueError("El proveedor seleccionado esta inactivo.")
        if expected_type and row["tipo"] != expected_type:
            raise ValueError("El proveedor no corresponde al tipo esperado.")

    def _get_tipo_gasto_flota(self, cur, tipo_gasto_id):
        cur.execute(
            """
            SELECT id, nombre, requiere_km, activo
            FROM tipos_gasto_flota
            WHERE id = %s
            """,
            (int(tipo_gasto_id),),
        )
        row = cur.fetchone()
        if not row or not bool(row["activo"]):
            raise ValueError("Tipo de gasto no disponible.")
        return row

    def _validate_factura_combustible_unica(self, cur, proveedor_id, nro_factura: str, exclude_id=None):
        normalized = _normalize_invoice_number(nro_factura)
        if not normalized:
            return
        filters = [
            "LOWER(REGEXP_REPLACE(BTRIM(COALESCE(nro_factura, '')), '[[:space:]]+', '', 'g')) = %s",
            "COALESCE(proveedor_id, 0) = %s",
            "eliminado_en IS NULL",
        ]
        if exclude_id:
            filters.append("id <> %s")
        cur.execute(
            f"""
            SELECT id
            FROM cargas_combustible
            WHERE {" AND ".join(filters)}
            LIMIT 1
            """,
            (normalized, int(proveedor_id or 0), int(exclude_id)) if exclude_id else (normalized, int(proveedor_id or 0)),
        )
        row = cur.fetchone()
        if row:
            raise ValueError(f"La factura {nro_factura} ya esta registrada para ese proveedor de combustible.")

    def _validate_factura_gasto_unica(self, cur, proveedor_id, proveedor_nombre: str, proveedor_ruc: str, nro_factura: str, exclude_id=None):
        normalized = _normalize_invoice_number(nro_factura)
        if not normalized:
            return
        provider_key = self._gasto_factura_provider_key(proveedor_id, proveedor_nombre, proveedor_ruc)
        filters = [
            "LOWER(REGEXP_REPLACE(BTRIM(COALESCE(nro_factura, '')), '[[:space:]]+', '', 'g')) = %s",
            """
            CASE
                WHEN proveedor_id IS NOT NULL THEN 'id:' || proveedor_id::text
                WHEN BTRIM(COALESCE(proveedor_ruc, '')) <> '' THEN 'ruc:' || LOWER(REGEXP_REPLACE(BTRIM(COALESCE(proveedor_ruc, '')), '[[:space:]]+', '', 'g'))
                ELSE 'nombre:' || LOWER(REGEXP_REPLACE(BTRIM(COALESCE(proveedor_nombre, '')), '[[:space:]]+', '', 'g'))
            END = %s
            """,
            "eliminado_en IS NULL",
        ]
        if exclude_id:
            filters.append("id <> %s")
        cur.execute(
            f"""
            SELECT id
            FROM gastos_flota
            WHERE {" AND ".join(filters)}
            LIMIT 1
            """,
            (normalized, provider_key, int(exclude_id)) if exclude_id else (normalized, provider_key),
        )
        row = cur.fetchone()
        if row:
            raise ValueError(f"La factura {nro_factura} ya esta registrada para ese proveedor en gastos de flota.")

    def _gasto_factura_provider_key(self, proveedor_id, proveedor_nombre: str, proveedor_ruc: str) -> str:
        if proveedor_id:
            return f"id:{int(proveedor_id)}"
        normalized_ruc = _normalize_invoice_number(proveedor_ruc)
        if normalized_ruc:
            return f"ruc:{normalized_ruc}"
        return f"nombre:{_normalize_invoice_number(proveedor_nombre)}"

    def _validate_km_flota(self, cur, vehiculo_id, km_actual):
        if km_actual is None:
            return
        cur.execute(
            """
            WITH historial AS (
                SELECT km_actual
                FROM cargas_combustible
                WHERE vehiculo_id = %s AND km_actual IS NOT NULL AND eliminado_en IS NULL
                UNION ALL
                SELECT km_actual
                FROM gastos_flota
                WHERE vehiculo_id = %s AND km_actual IS NOT NULL AND eliminado_en IS NULL
            )
            SELECT MAX(km_actual) AS ultimo_km
            FROM historial
            """,
            (int(vehiculo_id), int(vehiculo_id)),
        )
        row = cur.fetchone() or {}
        ultimo_km = row.get("ultimo_km")
        if ultimo_km is not None and float(km_actual) < float(ultimo_km):
            raise ValueError(f"El kilometraje no puede ser menor al ultimo registrado ({ultimo_km}).")

    def _read_combustible_import_rows(self, file_name: str, file_content: str):
        raw_base64 = file_content.split(",", 1)[1] if "," in file_content else file_content
        try:
            raw = base64.b64decode(raw_base64)
        except Exception as exc:
            raise ValueError("No se pudo leer el archivo adjunto.") from exc

        lower_name = file_name.lower()
        if lower_name.endswith(".csv"):
            text = None
            for encoding in ("utf-8-sig", "cp1252", "latin-1"):
                try:
                    text = raw.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            if text is None:
                raise ValueError("No se pudo leer el CSV. Verifica la codificacion del archivo.")
            lines = [line for line in text.splitlines() if line.strip()]
            if not lines:
                return []
            delimiter = ";"
            if sum(line.count(";") for line in lines[:5]) == 0 and sum(line.count(",") for line in lines[:5]) > 0:
                delimiter = ","
            header_idx = 0
            for index, line in enumerate(lines):
                normalized_cells = [_normalize_import_header(cell) for cell in line.split(delimiter)]
                score = len([cell for cell in normalized_cells if cell in {"vehiculo", "fecha real", "fecha", "producto", "cantidad", "litros", "precio", "importe", "codigo de autorizacion"}])
                if score >= 3:
                    header_idx = index
                    break
            csv_text = "\n".join(lines[header_idx:])
            reader = csv.DictReader(StringIO(csv_text), delimiter=delimiter)
            return [
                (header_idx + index + 1, {_normalize_import_header(key): value for key, value in (row or {}).items()})
                for index, row in enumerate(reader, start=1)
            ]

        if lower_name.endswith(".xlsx") or lower_name.endswith(".xlsm"):
            if openpyxl is None:
                raise ValueError("Falta soporte XLSX en el servidor.")
            wb = openpyxl.load_workbook(BytesIO(raw), data_only=True, read_only=True)
            expected = {
                "vehiculo",
                "fecha real",
                "fecha",
                "producto",
                "cantidad",
                "litros",
                "precio",
                "importe",
                "codigo de autorizacion",
            }
            best_sheet = None
            best_header_row = None
            best_headers = None
            best_score = -1
            for ws in wb.worksheets:
                for row_number, values in enumerate(ws.iter_rows(min_row=1, max_row=25, values_only=True), start=1):
                    headers = [_normalize_import_header(value) for value in (values or [])]
                    score = len([header for header in headers if header in expected])
                    if score > best_score:
                        best_score = score
                        best_sheet = ws
                        best_header_row = row_number
                        best_headers = headers
            if not best_sheet or not best_headers or best_score < 3:
                raise ValueError("No se encontraron encabezados validos en el Excel.")
            rows = []
            for row_number, values in enumerate(best_sheet.iter_rows(min_row=(best_header_row or 1) + 1, values_only=True), start=(best_header_row or 1) + 1):
                row = {best_headers[index]: values[index] for index in range(min(len(best_headers), len(values))) if best_headers[index]}
                rows.append((row_number, row))
            return rows

        raise ValueError("Formato no soportado. Usa CSV o XLSX.")

    def _map_combustible_import_row(self, row: dict[str, Any]):
        def pick(*keys):
            for key in keys:
                value = row.get(key)
                if value not in (None, ""):
                    return value
            return None

        vehiculo_ref = str(pick("vehiculo", "codigo", "codigo vehiculo", "movil", "nro movil") or "").strip()
        fecha = _parse_flexible_date(pick("fecha real", "fecha", "fecha carga"))
        tipo_combustible = str(pick("producto", "tipo combustible", "combustible") or "").strip()
        litros = _parse_number(pick("cantidad", "litros", "cantidad litros") or 0)
        importe_raw = pick("importe", "monto", "total")
        precio_raw = pick("precio", "precio litro", "costo litro")
        importe = _parse_number(importe_raw) if importe_raw not in (None, "") else 0
        precio = _parse_number(precio_raw) if precio_raw not in (None, "") else 0
        nro_factura = str(pick("codigo de autorizacion", "autorizacion", "factura", "nro factura") or "").strip()

        if not vehiculo_ref:
            raise ValueError("Falta la columna Vehiculo.")
        if not fecha:
            raise ValueError("La fecha no tiene un formato valido.")
        if litros <= 0:
            raise ValueError("La cantidad/litros debe ser mayor a 0.")
        if importe <= 0 and precio > 0:
            importe = round(litros * precio, 3)
        if importe <= 0:
            raise ValueError("La fila no tiene importe valido.")

        return {
            "vehiculo_ref": vehiculo_ref,
            "fecha": fecha,
            "tipo_combustible": tipo_combustible,
            "litros": litros,
            "importe": importe,
            "nro_factura": nro_factura,
            "observacion": "Importado desde archivo",
        }

    def _find_vehiculo_flota_import(self, cur, raw_ref: str):
        ref = str(raw_ref or "").strip()
        if not ref:
            raise ValueError("Referencia de vehiculo vacia.")
        ref_norm = _normalize_vehicle_import_ref(ref)
        if not ref_norm:
            raise ValueError("Referencia de vehiculo vacia.")
        cur.execute(
            """
            SELECT id, codigo, chapa, nombre, sucursal, activo
            FROM vehiculos
            ORDER BY activo DESC, id
            """,
        )
        rows = cur.fetchall()
        if not rows:
            raise ValueError(f"No existe un vehiculo que coincida con '{ref}'.")

        exact_matches = []
        partial_matches = []
        for row in rows:
            normalized_values = [
                _normalize_vehicle_import_ref(row.get("codigo")),
                _normalize_vehicle_import_ref(row.get("chapa")),
                _normalize_vehicle_import_ref(row.get("nombre")),
            ]
            normalized_values = [value for value in normalized_values if value]
            if any(value == ref_norm for value in normalized_values):
                exact_matches.append(row)
                continue
            if any(ref_norm in value or value in ref_norm for value in normalized_values):
                partial_matches.append(row)

        matches = exact_matches or partial_matches
        if not matches:
            best_match = None
            best_ratio = 0.0
            for row in rows:
                normalized_values = [
                    _normalize_vehicle_import_ref(row.get("codigo")),
                    _normalize_vehicle_import_ref(row.get("chapa")),
                    _normalize_vehicle_import_ref(row.get("nombre")),
                ]
                normalized_values = [value for value in normalized_values if value]
                for value in normalized_values:
                    ratio = difflib.SequenceMatcher(None, ref_norm, value).ratio()
                    if ratio > best_ratio:
                        best_ratio = ratio
                        best_match = row
            if best_match is not None and best_ratio >= 0.84:
                matches = [best_match]
        if not matches:
            raise ValueError(f"No existe un vehiculo que coincida con '{ref}'.")
        if len(matches) > 1:
            raise ValueError(f"Hay mas de un vehiculo que coincide con '{ref}'.")
        if not bool(matches[0]["activo"]):
            raise ValueError(f"El vehiculo '{ref}' esta inactivo.")
        return matches[0]

    def _vehiculo_flota_label(self, vehiculo: dict[str, Any]) -> str:
        codigo = str(vehiculo.get("codigo") or "").strip()
        chapa = str(vehiculo.get("chapa") or "").strip()
        nombre = str(vehiculo.get("nombre") or "").strip()
        principal = codigo or chapa or (f"Vehiculo #{vehiculo.get('id')}" if vehiculo.get("id") else "Vehiculo")
        if nombre and nombre.lower() != principal.lower():
            return f"{principal} - {nombre}"
        return nombre or principal

    def get_distribuciones_data(self, lote_id=None):
        with self._connect(readonly=True) as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """
                    SELECT
                        L.id,
                        L.lote,
                        TRIM(L.empresa) AS empresa,
                        L.fecha,
                        COALESCE(SUM(F.cantidad), 0)::int AS faenado,
                        (
                            SELECT COALESCE(SUM(D.cabezas), 0)::int
                            FROM distribuciones D
                            WHERE D.lote_id = L.id
                        ) AS distribuidas,
                        EXISTS (
                            SELECT 1
                            FROM distribuciones D0
                            WHERE D0.lote_id = L.id AND COALESCE(D0.kg, 0) = 0
                        ) AS has_zero_kg
                    FROM lotes L
                    LEFT JOIN faenas F ON F.lote_id = L.id
                    GROUP BY L.id, L.lote, L.empresa, L.fecha
                    HAVING COALESCE(SUM(F.cantidad), 0) > 0
                    ORDER BY L.fecha DESC, L.lote
                    """
                )
                lotes = cur.fetchall()

                selected_lote_id = int(lote_id) if lote_id else (int(lotes[0]["id"]) if lotes else None)
                distribuciones = []
                resumen_local = []
                if selected_lote_id:
                    cur.execute(
                        """
                        SELECT id,
                               fecha,
                               local,
                               kg,
                               COALESCE(nota, '') AS nota,
                               cabezas,
                               COALESCE(diferencia_kg, 0) AS diferencia_kg
                        FROM distribuciones
                        WHERE lote_id = %s
                        ORDER BY fecha DESC, id DESC
                        """,
                        (selected_lote_id,),
                    )
                    distribuciones = cur.fetchall()

                    cur.execute(
                        """
                        SELECT local,
                               COALESCE(SUM(kg), 0) AS kg,
                               COALESCE(SUM(cabezas), 0)::int AS cabezas
                        FROM distribuciones
                        WHERE lote_id = %s
                        GROUP BY local
                        ORDER BY local
                        """,
                        (selected_lote_id,),
                    )
                    resumen_local = cur.fetchall()

        return {
            "lotes": lotes,
            "selected_lote_id": selected_lote_id,
            "distribuciones": distribuciones,
            "resumenLocal": resumen_local,
        }

    def save_distribucion(self, payload):
        distrib_id = payload.get("id")
        lote_id = int(payload.get("lote_id"))
        fecha = str(payload.get("fecha") or "").strip()
        local = str(payload.get("local") or "").strip().upper()
        kg = _parse_number(payload.get("kg") or 0)
        cabezas = _parse_int(payload.get("cabezas") or 0)
        nota = str(payload.get("nota") or "").strip()
        diferencia_kg = _parse_number(payload.get("diferencia_kg") or 0)

        if not fecha:
            raise ValueError("Fecha requerida.")
        if local not in {"LUQUE", "AREGUA", "ITAUGUA"}:
            raise ValueError("Local invalido.")
        if kg < 0:
            raise ValueError("El kilaje (kg) debe ser mayor o igual a 0.")
        if cabezas < 0:
            raise ValueError("La cantidad de reces debe ser mayor o igual a 0.")

        with self._connect(readonly=False) as conn:
            try:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute("SELECT COALESCE(SUM(cantidad), 0)::int AS faenado FROM faenas WHERE lote_id = %s", (lote_id,))
                    faenado = int((cur.fetchone() or {}).get("faenado") or 0)
                    if faenado <= 0:
                        raise ValueError("El lote no tiene faena registrada.")

                    cur.execute("SELECT COALESCE(SUM(cabezas), 0)::int AS total FROM distribuciones WHERE lote_id = %s", (lote_id,))
                    total_cabezas = int((cur.fetchone() or {}).get("total") or 0)
                    old_cabezas = 0
                    if distrib_id:
                        cur.execute(
                            "SELECT cabezas FROM distribuciones WHERE id = %s AND lote_id = %s",
                            (int(distrib_id), lote_id),
                        )
                        row = cur.fetchone()
                        if not row:
                            raise ValueError("Distribucion no encontrada.")
                        old_cabezas = int(row["cabezas"] or 0)

                    otras_cabezas = total_cabezas - old_cabezas
                    if (otras_cabezas + cabezas) > faenado:
                        saldo = faenado - otras_cabezas
                        raise ValueError(f"No hay suficientes cabezas disponibles. Maximo permitido: {saldo}.")

                    if distrib_id:
                        cur.execute(
                            """
                            UPDATE distribuciones
                            SET fecha = %s, local = %s, kg = %s, cabezas = %s, nota = %s, diferencia_kg = %s
                            WHERE id = %s
                            RETURNING id, lote_id, fecha, local, kg, cabezas, nota, diferencia_kg
                            """,
                            (fecha, local, kg, cabezas, nota, diferencia_kg, int(distrib_id)),
                        )
                    else:
                        cur.execute(
                            """
                            INSERT INTO distribuciones(lote_id, fecha, local, kg, cabezas, nota, diferencia_kg)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                            RETURNING id, lote_id, fecha, local, kg, cabezas, nota, diferencia_kg
                            """,
                            (lote_id, fecha, local, kg, cabezas, nota, diferencia_kg),
                        )
                    row = cur.fetchone()
                conn.commit()
                return row
            except Exception:
                conn.rollback()
                raise

    def delete_distribucion(self, distrib_id):
        with self._connect(readonly=False) as conn:
            try:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute(
                        """
                        SELECT id, lote_id, fecha, local, kg, COALESCE(nota, '') AS nota, cabezas, COALESCE(diferencia_kg, 0) AS diferencia_kg
                        FROM distribuciones
                        WHERE id = %s
                        """,
                        (int(distrib_id),),
                    )
                    row = cur.fetchone()
                    if not row:
                        raise ValueError("Distribucion no encontrada.")
                    cur.execute(
                        """
                        INSERT INTO distribuciones_eliminadas(
                            distribucion_id, lote_id, fecha, local, kg, nota, cabezas, diferencia_kg
                        )
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                        """,
                        (
                            int(row["id"]),
                            int(row["lote_id"]),
                            row["fecha"],
                            row["local"],
                            float(row["kg"]),
                            row["nota"],
                            int(row["cabezas"]),
                            float(row["diferencia_kg"] or 0),
                        ),
                    )
                    cur.execute("DELETE FROM distribuciones WHERE id = %s", (int(distrib_id),))
                conn.commit()
            except Exception:
                conn.rollback()
                raise
        return {"ok": True}

    def get_compras_faena_data(self, lote_id=None):
        with self._connect(readonly=True) as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """
                    SELECT L.id,
                           L.lote,
                           TRIM(L.empresa) AS empresa,
                           L.fecha,
                           L.cantidad,
                           COALESCE(F.faenado, 0)::int AS faenado,
                           (L.cantidad - COALESCE(F.faenado, 0))::int AS restante,
                           COALESCE(D.distribuidas, 0)::int AS distribuidas,
                           L.monto,
                           COALESCE(L.peso_compra_kg, 0.0) AS peso_compra_kg,
                           COALESCE(L.cantidad_vac, 0)::int AS cantidad_vac,
                           COALESCE(L.cantidad_tor, 0)::int AS cantidad_tor,
                           COALESCE(L.cantidad_nov, 0)::int AS cantidad_nov,
                           COALESCE(L.cantidad_vaq, 0)::int AS cantidad_vaq,
                           COALESCE(L.peso_promedio_vac, 0.0) AS peso_promedio_vac,
                           COALESCE(L.peso_promedio_tor, 0.0) AS peso_promedio_tor,
                           COALESCE(L.peso_promedio_nov, 0.0) AS peso_promedio_nov,
                           COALESCE(L.peso_promedio_vaq, 0.0) AS peso_promedio_vaq,
                           COALESCE(L.cerrado, false) AS cerrado
                    FROM lotes L
                    LEFT JOIN (
                        SELECT lote_id, SUM(cantidad)::int AS faenado
                        FROM faenas
                        GROUP BY lote_id
                    ) F ON F.lote_id = L.id
                    LEFT JOIN (
                        SELECT lote_id, SUM(cabezas)::int AS distribuidas
                        FROM distribuciones
                        GROUP BY lote_id
                    ) D ON D.lote_id = L.id
                    ORDER BY L.fecha DESC, L.lote
                    """
                )
                lotes = cur.fetchall()
                selected_lote_id = int(lote_id) if lote_id else (int(lotes[0]["id"]) if lotes else None)

                pendientes = [row for row in lotes if int(row["restante"] or 0) > 0]
                completados = [row for row in lotes if int(row["restante"] or 0) <= 0]
                resumen = {
                    "lotes_registrados": len(lotes),
                    "reses_camara": 0,
                    "reses_sin_faenar": sum(max(int(row["restante"] or 0), 0) for row in lotes),
                    "lotes_pendientes": len(pendientes),
                }

                cur.execute(
                    """
                    SELECT COALESCE(SUM(GREATEST(f.faenado - COALESCE(d.distribuidas, 0), 0)), 0)::int AS reses_camara
                    FROM (
                        SELECT lote_id, SUM(cantidad)::int AS faenado
                        FROM faenas
                        GROUP BY lote_id
                    ) f
                    LEFT JOIN (
                        SELECT lote_id, SUM(cabezas)::int AS distribuidas
                        FROM distribuciones
                        GROUP BY lote_id
                    ) d ON d.lote_id = f.lote_id
                    """
                )
                resumen["reses_camara"] = int((cur.fetchone() or {}).get("reses_camara") or 0)

                faenas = []
                if selected_lote_id:
                    cur.execute(
                        """
                        SELECT id, lote_id, fecha, cantidad, COALESCE(nota, '') AS nota
                        FROM faenas
                        WHERE lote_id = %s
                        ORDER BY fecha ASC, id ASC
                        """,
                        (selected_lote_id,),
                    )
                    faenas = cur.fetchall()

        return {
            "empresas": EMPRESAS,
            "resumen": resumen,
            "lotes": lotes,
            "pendientes": pendientes,
            "completados": completados,
            "selected_lote_id": selected_lote_id,
            "faenas": faenas,
        }

    def save_lote(self, payload):
        lote_id = payload.get("id")
        lote = str(payload.get("lote") or "").strip()
        empresa = str(payload.get("empresa") or "").strip()
        fecha = str(payload.get("fecha") or "").strip()
        cantidad = _parse_int(payload.get("cantidad") or 0)
        monto = _parse_number(payload.get("monto") or 0)
        peso_compra_kg = _parse_number(payload.get("peso_compra_kg") or 0)
        cantidad_vac = _parse_int(payload.get("cantidad_vac") or 0)
        cantidad_tor = _parse_int(payload.get("cantidad_tor") or 0)
        cantidad_nov = _parse_int(payload.get("cantidad_nov") or 0)
        cantidad_vaq = _parse_int(payload.get("cantidad_vaq") or 0)
        peso_promedio_vac = _parse_number(payload.get("peso_promedio_vac") or 0)
        peso_promedio_tor = _parse_number(payload.get("peso_promedio_tor") or 0)
        peso_promedio_nov = _parse_number(payload.get("peso_promedio_nov") or 0)
        peso_promedio_vaq = _parse_number(payload.get("peso_promedio_vaq") or 0)
        clasificacion_total = cantidad_vac + cantidad_tor + cantidad_nov + cantidad_vaq

        if not lote or not empresa or not fecha:
            raise ValueError("Complete lote, empresa y fecha.")
        if empresa not in EMPRESAS:
            raise ValueError("Empresa invalida.")
        if cantidad <= 0:
            raise ValueError("La cantidad debe ser mayor a 0.")
        if monto < 0:
            raise ValueError("El monto no puede ser negativo.")
        if peso_compra_kg < 0:
            raise ValueError("El peso total debe ser mayor o igual a 0.")
        if any(value < 0 for value in (cantidad_vac, cantidad_tor, cantidad_nov, cantidad_vaq)):
            raise ValueError("Las cantidades por clase no pueden ser negativas.")
        if any(value < 0 for value in (peso_promedio_vac, peso_promedio_tor, peso_promedio_nov, peso_promedio_vaq)):
            raise ValueError("Los pesos promedio por clase no pueden ser negativos.")
        if clasificacion_total != cantidad:
            raise ValueError("La suma VAC + TOR + NOV + VAQ debe ser igual a la cantidad comprada.")

        with self._connect(readonly=False) as conn:
            try:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    if lote_id:
                        cur.execute(
                            "SELECT COALESCE(SUM(cantidad), 0)::int AS faenado FROM faenas WHERE lote_id = %s",
                            (int(lote_id),),
                        )
                        faenado = int((cur.fetchone() or {}).get("faenado") or 0)
                        if cantidad < faenado:
                            raise ValueError(f"No se puede establecer una cantidad menor a la faenada. Faenado actual: {faenado}.")
                        cur.execute(
                            """
                            UPDATE lotes
                            SET lote = %s,
                                empresa = %s,
                                fecha = %s,
                                cantidad = %s,
                                monto = %s,
                                peso_compra_kg = %s,
                                cantidad_vac = %s,
                                cantidad_tor = %s,
                                cantidad_nov = %s,
                                cantidad_vaq = %s,
                                peso_promedio_vac = %s,
                                peso_promedio_tor = %s,
                                peso_promedio_nov = %s,
                                peso_promedio_vaq = %s
                            WHERE id = %s
                            RETURNING id, lote, empresa, fecha, cantidad, monto,
                                      COALESCE(peso_compra_kg, 0.0) AS peso_compra_kg,
                                      cantidad_vac, cantidad_tor, cantidad_nov, cantidad_vaq,
                                      peso_promedio_vac, peso_promedio_tor, peso_promedio_nov, peso_promedio_vaq
                            """,
                            (
                                lote,
                                empresa,
                                fecha,
                                cantidad,
                                monto,
                                peso_compra_kg,
                                cantidad_vac,
                                cantidad_tor,
                                cantidad_nov,
                                cantidad_vaq,
                                peso_promedio_vac,
                                peso_promedio_tor,
                                peso_promedio_nov,
                                peso_promedio_vaq,
                                int(lote_id),
                            ),
                        )
                    else:
                        cur.execute(
                            """
                            INSERT INTO lotes(
                                lote, empresa, fecha, cantidad, monto, peso_compra_kg,
                                cantidad_vac, cantidad_tor, cantidad_nov, cantidad_vaq,
                                peso_promedio_vac, peso_promedio_tor, peso_promedio_nov, peso_promedio_vaq
                            )
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            RETURNING id, lote, empresa, fecha, cantidad, monto,
                                      COALESCE(peso_compra_kg, 0.0) AS peso_compra_kg,
                                      cantidad_vac, cantidad_tor, cantidad_nov, cantidad_vaq,
                                      peso_promedio_vac, peso_promedio_tor, peso_promedio_nov, peso_promedio_vaq
                            """,
                            (
                                lote,
                                empresa,
                                fecha,
                                cantidad,
                                monto,
                                peso_compra_kg,
                                cantidad_vac,
                                cantidad_tor,
                                cantidad_nov,
                                cantidad_vaq,
                                peso_promedio_vac,
                                peso_promedio_tor,
                                peso_promedio_nov,
                                peso_promedio_vaq,
                            ),
                        )
                    row = cur.fetchone()
                    if not row:
                        raise ValueError("Lote no encontrado.")
                conn.commit()
                return row
            except Exception:
                conn.rollback()
                raise

    def delete_lote_compra(self, payload):
        lote_id = _parse_int(payload.get("id") or payload.get("lote_id") or 0)
        if lote_id <= 0:
            raise ValueError("Falta el id del lote.")

        with self._connect(readonly=False) as conn:
            try:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute(
                        """
                        SELECT L.id,
                               L.lote,
                               COALESCE(SUM(F.cantidad), 0)::int AS faenado
                        FROM lotes L
                        LEFT JOIN faenas F ON F.lote_id = L.id
                        WHERE L.id = %s
                        GROUP BY L.id, L.lote
                        """,
                        (lote_id,),
                    )
                    lote = cur.fetchone()
                    if not lote:
                        raise ValueError("Lote no encontrado.")
                    if int(lote["faenado"] or 0) > 0:
                        raise ValueError("No se puede eliminar un lote que ya fue faenado.")

                    cur.execute("DELETE FROM distribuciones WHERE lote_id = %s", (lote_id,))
                    distribuciones = cur.rowcount
                    cur.execute("DELETE FROM faenas WHERE lote_id = %s", (lote_id,))
                    faenas = cur.rowcount
                    cur.execute("DELETE FROM lotes WHERE id = %s", (lote_id,))
                    if cur.rowcount != 1:
                        raise ValueError("No se pudo eliminar el lote.")
                conn.commit()
                return {"ok": True, "lote_id": lote_id, "lote": lote["lote"], "faenas_eliminadas": faenas, "distribuciones_eliminadas": distribuciones}
            except Exception:
                conn.rollback()
                raise

    def set_faena_total(self, payload):
        lote_id = int(payload.get("lote_id"))
        fecha = str(payload.get("fecha") or "").strip()
        cantidad_total = _parse_int(payload.get("cantidad_total") or 0)
        nota = str(payload.get("nota") or "").strip()

        if not fecha:
            raise ValueError("Fecha requerida.")
        if cantidad_total < 0:
            raise ValueError("La faena total debe ser mayor o igual a 0.")

        with self._connect(readonly=False) as conn:
            try:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute(
                        """
                        SELECT L.id,
                               L.cantidad::int AS compra,
                               COALESCE(D.distribuidas, 0)::int AS distribuidas
                        FROM lotes L
                        LEFT JOIN (
                            SELECT lote_id, SUM(cabezas)::int AS distribuidas
                            FROM distribuciones
                            GROUP BY lote_id
                        ) D ON D.lote_id = L.id
                        WHERE L.id = %s
                        """,
                        (lote_id,),
                    )
                    row = cur.fetchone()
                    if not row:
                        raise ValueError("Lote no encontrado.")

                    compra = int(row["compra"] or 0)
                    distribuidas = int(row["distribuidas"] or 0)
                    if cantidad_total > compra:
                        raise ValueError(f"La faena total ({cantidad_total}) no puede superar la compra ({compra}).")
                    if cantidad_total < distribuidas:
                        raise ValueError(
                            f"La faena total ({cantidad_total}) no puede ser menor a lo ya distribuido ({distribuidas})."
                        )

                    cur.execute("DELETE FROM faenas WHERE lote_id = %s", (lote_id,))
                    if cantidad_total > 0:
                        cur.execute(
                            """
                            INSERT INTO faenas(lote_id, fecha, cantidad, nota)
                            VALUES (%s, %s, %s, %s)
                            RETURNING id, lote_id, fecha, cantidad, COALESCE(nota, '') AS nota
                            """,
                            (lote_id, fecha, cantidad_total, nota),
                        )
                        saved = cur.fetchone()
                    else:
                        saved = {"id": None, "lote_id": lote_id, "fecha": fecha, "cantidad": 0, "nota": nota}
                conn.commit()
                return saved
            except Exception:
                conn.rollback()
                raise

    def add_faena(self, payload):
        lote_id = int(payload.get("lote_id"))
        fecha = str(payload.get("fecha") or "").strip()
        cantidad = _parse_int(payload.get("cantidad") or 0)
        nota = str(payload.get("nota") or "").strip()

        if not fecha:
            raise ValueError("Fecha requerida.")
        if cantidad <= 0:
            raise ValueError("La cantidad de faena debe ser mayor a 0.")

        with self._connect(readonly=False) as conn:
            try:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute(
                        """
                        SELECT L.cantidad - COALESCE(SUM(F.cantidad), 0)::int AS restante
                        FROM lotes L
                        LEFT JOIN faenas F ON F.lote_id = L.id
                        WHERE L.id = %s
                        GROUP BY L.id
                        """,
                        (lote_id,),
                    )
                    row = cur.fetchone()
                    if not row:
                        raise ValueError("Lote no encontrado.")
                    restante = int(row["restante"] or 0)
                    if cantidad > restante:
                        raise ValueError(f"La cantidad a faenar ({cantidad}) excede el restante ({restante}).")
                    cur.execute(
                        """
                        INSERT INTO faenas(lote_id, fecha, cantidad, nota)
                        VALUES (%s, %s, %s, %s)
                        RETURNING id, lote_id, fecha, cantidad, COALESCE(nota, '') AS nota
                        """,
                        (lote_id, fecha, cantidad, nota),
                    )
                    saved = cur.fetchone()
                conn.commit()
                return saved
            except Exception:
                conn.rollback()
                raise

    def get_resumenes_data(self, lote_ids=None):
        selected_ids = []
        if lote_ids:
            selected_ids = [int(item) for item in str(lote_ids).split(",") if str(item).strip()]

        with self._connect(readonly=True) as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    f"""
                    WITH resumen_lotes AS ({self._resumen_lotes_cte()})
                    SELECT id,
                           lote,
                           empresa,
                           fecha,
                           cerrado,
                           cantcompra,
                           faenado,
                           distribuido,
                           kg,
                           monto,
                           costokg,
                           pct_distribuido,
                           pct_restante,
                           kgcompra,
                           cantidad_tor,
                           cantidad_nov,
                           cantidad_vac,
                           cantidad_vaq,
                           peso_promedio_tor,
                           peso_promedio_nov,
                           peso_promedio_vac,
                           peso_promedio_vaq,
                           rend_pct
                    FROM resumen_lotes
                    ORDER BY fecha DESC, lote
                    """
                )
                lotes = cur.fetchall()

                resumen_sucursales = [
                    {"local": "LUQUE", "kg": 0, "cabezas": 0, "dif_kg": 0},
                    {"local": "AREGUA", "kg": 0, "cabezas": 0, "dif_kg": 0},
                    {"local": "ITAUGUA", "kg": 0, "cabezas": 0, "dif_kg": 0},
                ]
                distribuciones_detalle = []
                if selected_ids:
                    cur.execute(
                        """
                        SELECT local,
                               COALESCE(SUM(kg), 0)::numeric AS kg,
                               COALESCE(SUM(cabezas), 0)::int AS cabezas,
                               COALESCE(SUM(diferencia_kg), 0)::numeric AS dif_kg
                        FROM distribuciones
                        WHERE lote_id = ANY(%s)
                        GROUP BY local
                        ORDER BY local
                        """,
                        (selected_ids,),
                    )
                    rows_by_local = {row["local"]: row for row in cur.fetchall()}
                    resumen_sucursales = [
                        rows_by_local.get(row["local"], row)
                        for row in resumen_sucursales
                    ]
                    cur.execute(
                        """
                        SELECT d.id,
                               d.lote_id,
                               l.lote,
                               d.fecha,
                               d.local,
                               d.kg,
                               d.cabezas,
                               COALESCE(d.diferencia_kg, 0) AS diferencia_kg,
                               COALESCE(d.nota, '') AS nota
                        FROM distribuciones d
                        JOIN lotes l ON l.id = d.lote_id
                        WHERE d.lote_id = ANY(%s)
                        ORDER BY l.fecha DESC, l.lote, d.fecha DESC, d.id DESC
                        """,
                        (selected_ids,),
                    )
                    distribuciones_detalle = cur.fetchall()

        return {
            "empresas": EMPRESAS,
            "lotes": lotes,
            "selected_lote_ids": selected_ids,
            "resumenSucursales": resumen_sucursales,
            "distribucionesDetalle": distribuciones_detalle,
        }

    def get_estadisticas_data(self, desde=None, hasta=None):
        with self._connect(readonly=True) as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                params = (desde, desde, hasta, hasta)
                cur.execute(
                    f"""
                    WITH resumen_lotes AS ({self._resumen_lotes_cte()}),
                    compras AS (
                        SELECT COUNT(*)::int AS lotes,
                               COALESCE(SUM(cantidad), 0)::int AS reces_compradas,
                               COALESCE(SUM(peso_compra_kg), 0)::numeric AS kg_compra_periodo,
                               COALESCE(SUM(monto), 0)::numeric AS monto_total
                        FROM lotes
                        WHERE (%s::date IS NULL OR fecha >= %s::date)
                          AND (%s::date IS NULL OR fecha <= %s::date)
                    ),
                    faena_periodo AS (
                        SELECT COALESCE(SUM(cantidad), 0)::int AS reces_faenadas
                        FROM faenas
                        WHERE (%s::date IS NULL OR fecha >= %s::date)
                          AND (%s::date IS NULL OR fecha <= %s::date)
                    ),
                    dist_lotes AS (
                        SELECT lote_id,
                               COALESCE(SUM(kg), 0)::numeric AS kg_distribuidos,
                               COALESCE(SUM(cabezas), 0)::int AS reces_distribuidas
                        FROM distribuciones
                        WHERE (%s::date IS NULL OR fecha >= %s::date)
                          AND (%s::date IS NULL OR fecha <= %s::date)
                        GROUP BY lote_id
                    ),
                    dist_periodo AS (
                        SELECT COALESCE(SUM(kg), 0)::numeric AS kg_distribuidos,
                               COALESCE(SUM(cabezas), 0)::int AS reces_distribuidas,
                               COUNT(DISTINCT lote_id) FILTER (WHERE COALESCE(kg, 0) = 0 AND COALESCE(cabezas, 0) > 0)::int AS lotes_kg_cero
                        FROM distribuciones
                        WHERE (%s::date IS NULL OR fecha >= %s::date)
                          AND (%s::date IS NULL OR fecha <= %s::date)
                    ),
                    dist_costos AS (
                        SELECT COALESCE(SUM(dl.kg_distribuidos * rl.costokg), 0)::numeric AS costo_ponderado,
                               COALESCE(SUM(dl.kg_distribuidos), 0)::numeric AS kg_distribuidos,
                               COALESCE(SUM(rl.kgcompra), 0)::numeric AS kg_compra_operada
                        FROM dist_lotes dl
                        JOIN resumen_lotes rl ON rl.id = dl.lote_id
                    ),
                    pendientes AS (
                        SELECT COUNT(*) FILTER (WHERE faenado > 0 AND distribuido < faenado)::int AS lotes_pendientes,
                               COUNT(*) FILTER (WHERE faenado > 0 AND distribuido = faenado)::int AS lotes_completados
                        FROM resumen_lotes
                    )
                    SELECT compras.lotes,
                           compras.reces_compradas,
                           faena_periodo.reces_faenadas,
                           dist_periodo.reces_distribuidas,
                           dist_periodo.kg_distribuidos,
                           dist_costos.kg_compra_operada AS kg_compra,
                           compras.monto_total,
                           CASE WHEN dist_costos.kg_distribuidos > 0
                                THEN ROUND((dist_costos.costo_ponderado / dist_costos.kg_distribuidos)::numeric, 2)
                                ELSE 0 END AS costo_kg_promedio,
                           CASE WHEN dist_costos.kg_compra_operada > 0
                                THEN ROUND((dist_costos.kg_distribuidos / dist_costos.kg_compra_operada) * 100, 2)
                                ELSE 0 END AS rendimiento_promedio,
                           pendientes.lotes_pendientes,
                           pendientes.lotes_completados,
                           dist_periodo.lotes_kg_cero
                    FROM compras, faena_periodo, dist_periodo, dist_costos, pendientes
                    """,
                    params + params + params + params,
                )
                kpis = cur.fetchone() or {}

                cur.execute(
                    f"""
                    WITH resumen_lotes AS ({self._resumen_lotes_cte()}),
                    dist_lotes AS (
                        SELECT lote_id,
                               COALESCE(SUM(kg), 0)::numeric AS kg_distribuidos,
                               COALESCE(SUM(cabezas), 0)::int AS reces_distribuidas
                        FROM distribuciones
                        WHERE (%s::date IS NULL OR fecha >= %s::date)
                          AND (%s::date IS NULL OR fecha <= %s::date)
                        GROUP BY lote_id
                    ),
                    base AS (
                        SELECT rl.*, dl.kg_distribuidos AS kg_periodo, dl.reces_distribuidas AS reces_periodo
                        FROM dist_lotes dl
                        JOIN resumen_lotes rl ON rl.id = dl.lote_id
                    )
                    SELECT empresa,
                           COUNT(*)::int AS lotes,
                           COALESCE(SUM(cantcompra), 0)::int AS reces_compradas,
                           COALESCE(SUM(kg_periodo), 0)::numeric AS kg_distribuidos,
                           COALESCE(SUM(kg_periodo * costokg), 0)::numeric AS monto_total,
                           CASE WHEN COALESCE(SUM(kg_periodo), 0) > 0
                                THEN ROUND((SUM(kg_periodo * costokg) / SUM(kg_periodo))::numeric, 2)
                                ELSE 0 END AS costo_kg,
                           CASE WHEN COALESCE(SUM(kgcompra), 0) > 0
                                THEN ROUND((SUM(kg_periodo)::numeric / SUM(kgcompra)) * 100, 2)
                                ELSE 0 END AS rendimiento,
                           CASE WHEN COALESCE((SELECT SUM(kg_periodo) FROM base), 0) > 0
                                THEN ROUND((SUM(kg_periodo)::numeric / (SELECT SUM(kg_periodo) FROM base)) * 100, 2)
                                ELSE 0 END AS participacion_pct
                    FROM base
                    GROUP BY empresa
                    ORDER BY kg_distribuidos DESC, monto_total DESC
                    """,
                    params,
                )
                proveedores = cur.fetchall()

                cur.execute(
                    """
                    SELECT local,
                           COALESCE(SUM(cabezas), 0)::int AS reces,
                           COALESCE(SUM(kg), 0)::numeric AS kg,
                           COALESCE(SUM(diferencia_kg), 0)::numeric AS dif_kg,
                           COUNT(*) FILTER (WHERE COALESCE(kg, 0) = 0)::int AS filas_kg_cero,
                           CASE WHEN COALESCE(SUM(kg), 0) > 0
                                THEN ROUND((ABS(COALESCE(SUM(diferencia_kg), 0)) / SUM(kg))::numeric * 100, 2)
                                ELSE 0 END AS desvio_pct,
                           CASE WHEN COALESCE((SELECT SUM(kg)
                                                FROM distribuciones
                                                WHERE (%s::date IS NULL OR fecha >= %s::date)
                                                  AND (%s::date IS NULL OR fecha <= %s::date)), 0) > 0
                                THEN ROUND((SUM(kg)::numeric / (SELECT SUM(kg)
                                                               FROM distribuciones
                                                               WHERE (%s::date IS NULL OR fecha >= %s::date)
                                                                 AND (%s::date IS NULL OR fecha <= %s::date))) * 100, 2)
                                ELSE 0 END AS participacion_pct
                    FROM distribuciones
                    WHERE (%s::date IS NULL OR fecha >= %s::date)
                      AND (%s::date IS NULL OR fecha <= %s::date)
                    GROUP BY local
                    ORDER BY ABS(COALESCE(SUM(diferencia_kg), 0)) DESC, local
                    """,
                    params + params + params,
                )
                sucursales = cur.fetchall()

                cur.execute(
                    """
                    WITH tipos AS (
                        SELECT 'TOR' AS tipo, COALESCE(SUM(cantidad_tor), 0)::int AS cantidad
                        FROM lotes
                        WHERE (%s::date IS NULL OR fecha >= %s::date)
                          AND (%s::date IS NULL OR fecha <= %s::date)
                        UNION ALL
                        SELECT 'NOV' AS tipo, COALESCE(SUM(cantidad_nov), 0)::int AS cantidad
                        FROM lotes
                        WHERE (%s::date IS NULL OR fecha >= %s::date)
                          AND (%s::date IS NULL OR fecha <= %s::date)
                        UNION ALL
                        SELECT 'VAC' AS tipo, COALESCE(SUM(cantidad_vac), 0)::int AS cantidad
                        FROM lotes
                        WHERE (%s::date IS NULL OR fecha >= %s::date)
                          AND (%s::date IS NULL OR fecha <= %s::date)
                        UNION ALL
                        SELECT 'VAQ' AS tipo, COALESCE(SUM(cantidad_vaq), 0)::int AS cantidad
                        FROM lotes
                        WHERE (%s::date IS NULL OR fecha >= %s::date)
                          AND (%s::date IS NULL OR fecha <= %s::date)
                    ),
                    total AS (
                        SELECT COALESCE(SUM(cantidad), 0)::numeric AS total_cantidad
                        FROM tipos
                    )
                    SELECT tipo,
                           cantidad,
                           CASE WHEN total.total_cantidad > 0
                                THEN ROUND((cantidad::numeric / total.total_cantidad) * 100, 2)
                                ELSE 0 END AS participacion_pct
                    FROM tipos, total
                    ORDER BY CASE tipo WHEN 'TOR' THEN 1 WHEN 'NOV' THEN 2 WHEN 'VAC' THEN 3 ELSE 4 END
                    """,
                    params + params + params + params,
                )
                clasificacion_compras = cur.fetchall()

                cur.execute(
                    f"""
                    WITH resumen_lotes AS ({self._resumen_lotes_cte()}),
                    dist_lotes AS (
                        SELECT lote_id, COALESCE(SUM(kg), 0)::numeric AS kg_periodo
                        FROM distribuciones
                        WHERE (%s::date IS NULL OR fecha >= %s::date)
                          AND (%s::date IS NULL OR fecha <= %s::date)
                        GROUP BY lote_id
                    )
                    SELECT rl.id, rl.lote, rl.empresa, rl.fecha, rl.faenado, rl.distribuido,
                           dl.kg_periodo AS kg, rl.kgcompra, rl.monto, rl.costokg, rl.rend_pct
                    FROM dist_lotes dl
                    JOIN resumen_lotes rl ON rl.id = dl.lote_id
                    WHERE dl.kg_periodo > 0
                    ORDER BY rl.rend_pct DESC, dl.kg_periodo DESC
                    LIMIT 10
                    """,
                    params,
                )
                mejores_lotes = cur.fetchall()

                cur.execute(
                    f"""
                    WITH resumen_lotes AS ({self._resumen_lotes_cte()})
                    SELECT id, lote, empresa, fecha, faenado, distribuido, kg, kgcompra, monto, costokg, rend_pct,
                           GREATEST(faenado - distribuido, 0)::int AS reces_pendientes
                    FROM resumen_lotes
                    WHERE faenado > 0 AND distribuido < faenado
                    ORDER BY fecha DESC, reces_pendientes DESC
                    LIMIT 10
                    """,
                )
                alertas = cur.fetchall()

                cur.execute(
                    f"""
                    WITH resumen_lotes AS ({self._resumen_lotes_cte()}),
                    actual_dist AS (
                        SELECT lote_id, COALESCE(SUM(kg), 0)::numeric AS kg_periodo
                        FROM distribuciones
                        WHERE (%s::date IS NULL OR fecha >= %s::date)
                          AND (%s::date IS NULL OR fecha <= %s::date)
                        GROUP BY lote_id
                    ),
                    actual AS (
                        SELECT rl.*, ad.kg_periodo
                        FROM actual_dist ad
                        JOIN resumen_lotes rl ON rl.id = ad.lote_id
                    ),
                    anterior_dist AS (
                        SELECT lote_id, COALESCE(SUM(kg), 0)::numeric AS kg_periodo
                        FROM distribuciones
                        WHERE fecha >= (COALESCE(%s::date, CURRENT_DATE) - INTERVAL '30 days')::date
                          AND fecha < COALESCE(%s::date, CURRENT_DATE)
                        GROUP BY lote_id
                    ),
                    anterior AS (
                        SELECT rl.*, ad.kg_periodo
                        FROM anterior_dist ad
                        JOIN resumen_lotes rl ON rl.id = ad.lote_id
                    ),
                    costo AS (
                        SELECT
                            CASE WHEN COALESCE((SELECT SUM(kg_periodo) FROM actual), 0) > 0
                                 THEN ((SELECT SUM(kg_periodo * costokg) FROM actual) / (SELECT SUM(kg_periodo) FROM actual))::numeric
                                 ELSE 0 END AS costo_actual,
                            CASE WHEN COALESCE((SELECT SUM(kg_periodo) FROM anterior), 0) > 0
                                 THEN ((SELECT SUM(kg_periodo * costokg) FROM anterior) / (SELECT SUM(kg_periodo) FROM anterior))::numeric
                                 ELSE 0 END AS costo_anterior
                    ),
                    rendimiento_bajo AS (
                        SELECT COUNT(*)::int AS cantidad
                        FROM actual
                        WHERE kg > 0 AND rend_pct > 0 AND rend_pct < 50
                    ),
                    pendientes AS (
                        SELECT COUNT(*)::int AS cantidad
                        FROM resumen_lotes
                        WHERE faenado > 0
                          AND distribuido < faenado
                          AND fecha <= CURRENT_DATE - INTERVAL '7 days'
                    ),
                    concentrado AS (
                        SELECT empresa,
                               CASE WHEN COALESCE((SELECT SUM(kg_periodo) FROM actual), 0) > 0
                                    THEN ROUND((SUM(kg_periodo)::numeric / (SELECT SUM(kg_periodo) FROM actual)) * 100, 2)
                                    ELSE 0 END AS participacion
                        FROM actual
                        GROUP BY empresa
                        ORDER BY participacion DESC
                        LIMIT 1
                    )
                    SELECT 'costo_kg_alto' AS tipo,
                           CASE
                               WHEN costo_anterior > 0 AND ((costo_actual - costo_anterior) / costo_anterior) * 100 >= 15 THEN 'alta'
                               WHEN costo_anterior > 0 AND ((costo_actual - costo_anterior) / costo_anterior) * 100 >= 8 THEN 'media'
                               ELSE 'baja'
                           END AS severidad,
                           'Costo/kg subio vs los 30 dias anteriores' AS titulo,
                           CASE
                               WHEN costo_anterior > 0 THEN ROUND(((costo_actual - costo_anterior) / costo_anterior) * 100, 2)
                               ELSE 0
                           END AS valor,
                           'Variacion porcentual del costo/kg promedio.' AS detalle
                    FROM costo
                    WHERE costo_anterior > 0 AND costo_actual > costo_anterior
                    UNION ALL
                    SELECT 'rendimiento_bajo',
                           CASE WHEN cantidad >= 5 THEN 'alta' WHEN cantidad >= 2 THEN 'media' ELSE 'baja' END,
                           'Lotes con rendimiento bajo',
                           cantidad,
                           'Cantidad de lotes con rendimiento menor a 50%%.'
                    FROM rendimiento_bajo
                    WHERE cantidad > 0
                    UNION ALL
                    SELECT 'proveedor_concentrado',
                           CASE WHEN participacion >= 60 THEN 'alta' WHEN participacion >= 45 THEN 'media' ELSE 'baja' END,
                           'Proveedor concentrado: ' || empresa,
                           participacion,
                           'Participacion sobre kg distribuidos del periodo.'
                    FROM concentrado
                    WHERE participacion >= 45
                    UNION ALL
                    SELECT 'pendiente_antiguo',
                           CASE WHEN cantidad >= 5 THEN 'alta' WHEN cantidad >= 2 THEN 'media' ELSE 'baja' END,
                           'Lotes pendientes hace mas de 7 dias',
                           cantidad,
                           'Lotes faenados con reces todavia no distribuidas.'
                    FROM pendientes
                    WHERE cantidad > 0
                    ORDER BY severidad DESC, tipo
                    """,
                    params + (hasta, hasta),
                )
                alertas_gestion = cur.fetchall()

        return {
            "kpis": kpis,
            "proveedores": proveedores,
            "sucursales": sucursales,
            "mejoresLotes": mejores_lotes,
            "alertas": alertas,
            "alertasGestion": alertas_gestion,
            "clasificacionCompras": clasificacion_compras,
        }

    def set_lotes_resumen_cerrado(self, payload):
        lote_ids = payload.get("lote_ids") or []
        cerrado = bool(payload.get("cerrado", True))

        if not isinstance(lote_ids, list):
            raise ValueError("lote_ids debe ser una lista.")

        ids = []
        for lote_id in lote_ids:
            value = _parse_int(lote_id)
            if value <= 0:
                continue
            ids.append(value)
        ids = sorted(set(ids))

        if not ids:
            raise ValueError("Selecciona al menos un lote.")

        with self._connect(readonly=False) as conn:
            try:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute(
                        """
                        UPDATE lotes
                        SET cerrado = %s
                        WHERE id = ANY(%s)
                        RETURNING id, COALESCE(cerrado, false) AS cerrado
                        """,
                        (cerrado, ids),
                    )
                    rows = cur.fetchall()
                    if not rows:
                        raise ValueError("No se encontraron lotes para actualizar.")
                conn.commit()
                return {"ok": True, "lotes": rows}
            except Exception:
                conn.rollback()
                raise

    def _pdf_wrap_cell_typewriter(self, text: str, align="LEFT", font_size=7, leading=8):
        align_map = {"LEFT": 0, "CENTER": 1, "RIGHT": 2, "JUSTIFY": 4}
        style = ParagraphStyle(
            name="PdfWrapCellTypewriter",
            fontName="Courier",
            fontSize=font_size,
            leading=leading,
            alignment=align_map.get(align, 0),
            spaceBefore=0,
            spaceAfter=0,
            wordWrap="CJK",
        )
        return Paragraph(escape(str(text or "")), style)

    def _build_table_compact_typewriter(self, data, col_widths=None, header_fill=None):
        header_fill = header_fill or colors.HexColor("#E9ECEF")
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), header_fill),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Courier-Bold"),
                    ("FONTNAME", (0, 1), (-1, -1), "Courier"),
                    ("FONTSIZE", (0, 0), (-1, 0), 7),
                    ("FONTSIZE", (0, 1), (-1, -1), 7),
                    ("LEADING", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
                ]
            )
        )
        return table

    def _build_pdf_metric_card(self, title: str, value: str, width=110, height=64):
        title_style = ParagraphStyle(
            name="PdfMetricCardTitle",
            fontName="Courier-Bold",
            fontSize=8,
            leading=9,
            alignment=1,
            textColor=colors.HexColor("#243447"),
            spaceBefore=0,
            spaceAfter=2,
        )
        value_style = ParagraphStyle(
            name="PdfMetricCardValue",
            fontName="Courier-Bold",
            fontSize=16,
            leading=18,
            alignment=1,
            textColor=colors.HexColor("#1F2937"),
            spaceBefore=0,
            spaceAfter=0,
        )
        card = Table(
            [[Paragraph(escape(title), title_style)], [Paragraph(escape(value), value_style)]],
            colWidths=[width],
            rowHeights=[24, max(height - 24, 24)],
        )
        card.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFF3CD")),
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#D6B656")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        return card

    def build_resumenes_pdf(self, lote_ids):
        if SimpleDocTemplate is None:
            raise RuntimeError("ReportLab no esta instalado. Instale reportlab para generar PDF.")
        selected_ids = [int(item) for item in str(lote_ids or "").split(",") if str(item).strip()]
        if not selected_ids:
            raise ValueError("Seleccione al menos un lote para generar el reporte.")

        with self._connect(readonly=True) as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    f"""
                    WITH resumen_lotes AS ({self._resumen_lotes_cte()})
                    SELECT *
                    FROM resumen_lotes
                    WHERE id = ANY(%s)
                    ORDER BY fecha DESC, lote
                    """,
                    (selected_ids,),
                )
                selected_rows = cur.fetchall()
                if not selected_rows:
                    raise ValueError("No se encontraron lotes seleccionados.")

                cur.execute(
                    """
                    WITH saldos AS (
                        SELECT L.id,
                               L.cantidad - COALESCE(SUM(F.cantidad), 0) AS restante
                        FROM lotes L
                        LEFT JOIN faenas F ON F.lote_id = L.id
                        GROUP BY L.id, L.cantidad
                    ),
                    camara AS (
                        SELECT L.id,
                               COALESCE(SUM(F.cantidad), 0) AS faenado,
                               (
                                   SELECT COALESCE(SUM(D.cabezas), 0)
                                   FROM distribuciones D
                                   WHERE D.lote_id = L.id
                               ) AS distribuidas
                        FROM lotes L
                        LEFT JOIN faenas F ON F.lote_id = L.id
                        GROUP BY L.id
                        HAVING COALESCE(SUM(F.cantidad), 0) > 0
                    )
                    SELECT
                        COALESCE((SELECT SUM(GREATEST(faenado - distribuidas, 0)) FROM camara), 0)::int AS total_camara,
                        COALESCE((SELECT SUM(GREATEST(restante, 0)) FROM saldos), 0)::int AS total_sin_faenar,
                        COALESCE((SELECT COUNT(*) FROM saldos WHERE restante > 0), 0)::int AS total_lotes_pend
                    """
                )
                globales = cur.fetchone() or {}

                dist_by_lote = {}
                for row in selected_rows:
                    cur.execute(
                        """
                        SELECT id, fecha, local, kg, COALESCE(nota, '') AS nota, cabezas, COALESCE(diferencia_kg, 0) AS dif_kg
                        FROM distribuciones
                        WHERE lote_id = %s
                        ORDER BY fecha DESC, id DESC
                        """,
                        (row["id"],),
                    )
                    dist_by_lote[int(row["id"])] = cur.fetchall()

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ReporteTypeTitle",
            parent=styles["Title"],
            fontName="Courier-Bold",
            fontSize=14,
            leading=16,
            spaceAfter=4,
        )
        heading_style = ParagraphStyle(
            "ReporteTypeHeading",
            parent=styles["Heading3"],
            fontName="Courier-Bold",
            fontSize=9,
            leading=10,
            spaceBefore=2,
            spaceAfter=4,
        )
        body_style = ParagraphStyle(
            "ReporteTypeBody",
            parent=styles["Normal"],
            fontName="Courier",
            fontSize=8,
            leading=9,
            spaceBefore=0,
            spaceAfter=2,
        )
        kpi_value_style = ParagraphStyle(
            "ReporteTypeKpiValue",
            parent=styles["Normal"],
            fontName="Courier-Bold",
            fontSize=14,
            leading=15,
            alignment=1,
            spaceBefore=0,
            spaceAfter=0,
        )

        total_monto_sel = sum(float(row["monto"] or 0) for row in selected_rows)
        total_kg_sel = sum(float(row["kg"] or 0) for row in selected_rows)
        costo_kg_sel = (total_monto_sel / total_kg_sel) if total_kg_sel > 0 else 0.0
        kpi_data = [[
            self._pdf_wrap_cell_typewriter("Reses en Camara", align="CENTER", font_size=8, leading=9),
            self._pdf_wrap_cell_typewriter("Reses sin Faenar", align="CENTER", font_size=8, leading=9),
            self._pdf_wrap_cell_typewriter("Lotes pendientes", align="CENTER", font_size=8, leading=9),
            self._pdf_wrap_cell_typewriter("Costo/kg reporte", align="CENTER", font_size=8, leading=9),
        ], [
            Paragraph(_fmt_int(globales.get("total_camara", 0)), kpi_value_style),
            Paragraph(_fmt_int(globales.get("total_sin_faenar", 0)), kpi_value_style),
            Paragraph(_fmt_int(globales.get("total_lotes_pend", 0)), kpi_value_style),
            Paragraph(_fmt_float(costo_kg_sel, 2), kpi_value_style),
        ]]
        story = [
            Paragraph("<b>Reporte de lotes completados</b>", title_style),
            Paragraph(
                f"Generado el <b>{datetime.now().strftime('%Y-%m-%d %H:%M')}</b> | Lotes incluidos: <b>{len(selected_rows)}</b>",
                body_style,
            ),
            Spacer(0, 6),
            self._build_table_compact_typewriter(kpi_data, col_widths=[170, 170, 150, 130], header_fill=colors.HexColor("#FFF3CD")),
            Spacer(0, 8),
        ]

        for row in selected_rows:
            story.append(Paragraph(f"<b>{row['empresa']}</b> | Lote <b>{row['lote']}</b>", heading_style))
            story.append(
                Paragraph(
                    (
                        f"Fecha: <b>{row['fecha']}</b> | Compra: <b>{_fmt_int(row['cantcompra'])}</b> | "
                        f"Faenado: <b>{_fmt_int(row['faenado'])}</b> | Distribuido: <b>{_fmt_int(row['distribuido'])}</b>"
                    ),
                    body_style,
                )
            )
            story.append(
                Paragraph(
                    (
                        f"Kg compra: <b>{_fmt_float(row['kgcompra'], 2)}</b> | "
                        f"Kg distribuidos: <b>{_fmt_float(row['kg'], 2)}</b> | "
                        f"Monto: <b>{_fmt_float(row['monto'], 0)}</b> | "
                        f"Rendimiento: <b>{_fmt_float(row['rend_pct'], 2)}%</b> | "
                        f"% distribuido: <b>{_fmt_float(row['pct_distribuido'], 2)}%</b>"
                    ),
                    body_style,
                )
            )
            story.append(Spacer(0, 4))

            data_dist = [["Fecha", "Sucursal", "Kg", "Dif. Kg", "Cantidad", "Nota"]]
            for dist in dist_by_lote.get(int(row["id"]), []):
                data_dist.append([
                    str(dist["fecha"]),
                    dist["local"],
                    _fmt_float(dist["kg"], 2),
                    _fmt_float(dist["dif_kg"], 2),
                    _fmt_int(dist["cabezas"]),
                    self._pdf_wrap_cell_typewriter(dist["nota"] or "", align="LEFT", font_size=7, leading=8),
                ])
            if len(data_dist) == 1:
                data_dist.append(["-", "-", "0,00", "0,00", "0", "Sin distribuciones"])

            story.append(Paragraph("<b>Recepciones / distribuciones registradas</b>", heading_style))
            dist_table = self._build_table_compact_typewriter(data_dist, col_widths=[58, 70, 54, 54, 54, 210])
            costo_card = self._build_pdf_metric_card("Costo/kg", _fmt_float(row["costokg"], 2), width=120, height=76)
            bloque_detalle = Table([[costo_card, dist_table]], colWidths=[130, 500])
            bloque_detalle.setStyle(
                TableStyle([
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ])
            )
            story.append(bloque_detalle)
            story.append(Spacer(0, 11))

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=18,
            rightMargin=18,
            topMargin=18,
            bottomMargin=18,
        )
        doc.build(story)
        return buffer.getvalue()


class DashboardHandler(BaseHTTPRequestHandler):
    repo = DashboardRepository(DATABASE_URL)
    auth = AuthService(DATABASE_URL)
    module_routes = discover_routes()

    def do_OPTIONS(self):
        self._send_empty(204)

    def do_GET(self):
        parsed = urlparse(self.path)
        query = parse_qs(parsed.query)
        if parsed.path == "/api/contratos":
            try:
                from web.backend.modules.contratos.repository import ContratosRepository

                self._require_module_access("contratos")
                return self._send_json(
                    ContratosRepository(DATABASE_URL).list_contratos(search=query.get("search", [""])[0])
                )
            except ValueError as exc:
                return self._send_json({"error": str(exc)}, status=400)
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)
        if parsed.path == "/api/acuerdos-comerciales/estadisticas":
            try:
                from web.backend.modules.acuerdos_comerciales.repository import AcuerdosComercialesRepository
                from web.backend.modules.acuerdos_comerciales.schemas import ROLES_ACUERDOS

                user = self._require_roles(ROLES_ACUERDOS)
                modules = user.get("modulos_permitidos")
                if isinstance(modules, list) and "acuerdos-estadisticas" not in modules:
                    raise PermissionDenied("No tienes permisos para ver estadisticas de acuerdos.")
                return self._send_json(
                    AcuerdosComercialesRepository(DATABASE_URL).get_estadisticas(
                        mes=query.get("mes", [None])[0],
                        anho=query.get("anho", [None])[0],
                    )
                )
            except ValueError as exc:
                return self._send_json({"error": str(exc)}, status=400)
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)
        if parsed.path == "/api/acuerdos-comerciales/cobranzas":
            try:
                from web.backend.modules.acuerdos_comerciales.repository import AcuerdosComercialesRepository
                from web.backend.modules.acuerdos_comerciales.schemas import ROLES_ACUERDOS

                self._require_roles(ROLES_ACUERDOS)
                mes = query.get("mes", [None])[0]
                anho = query.get("anho", [None])[0]
                if not mes or not anho:
                    raise ValueError("Mes y anho son obligatorios.")
                return self._send_json(AcuerdosComercialesRepository(DATABASE_URL).list_acuerdos_cobranzas(mes=mes, anho=anho))
            except ValueError as exc:
                return self._send_json({"error": str(exc)}, status=400)
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)
        if parsed.path == "/api/acuerdos-comerciales/cobranzas/anual":
            try:
                from web.backend.modules.acuerdos_comerciales.repository import AcuerdosComercialesRepository
                from web.backend.modules.acuerdos_comerciales.schemas import ROLES_ACUERDOS

                self._require_roles(ROLES_ACUERDOS)
                anho = query.get("anho", [None])[0]
                if not anho:
                    raise ValueError("Anho es obligatorio.")
                return self._send_json(AcuerdosComercialesRepository(DATABASE_URL).list_acuerdos_cobranzas_anual(anho=anho))
            except ValueError as exc:
                return self._send_json({"error": str(exc)}, status=400)
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)
        if parsed.path.rstrip("/") == "/api/dashboard/menudencias/pdf":
            try:
                user = self._require_roles({"admin", "supervisor"})
                desde = _parse_date(query.get("desde", [None])[0])
                hasta = _parse_date(query.get("hasta", [None])[0])
                pdf_bytes, filename = self.repo.build_menudencias_pdf(
                    desde=desde,
                    hasta=hasta,
                    generated_by=user.get("nombre") or user.get("username"),
                )
                return self._send_pdf(pdf_bytes, filename)
            except ValueError:
                return self._send_json({"error": "Formato de fecha invalido. Use YYYY-MM-DD."}, status=400)
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)
        if self._dispatch_module_route("GET", parsed, query):
            return
        if parsed.path == "/api/health":
            return self._send_json({"ok": True})
        if parsed.path == "/api/auth/me":
            user = self._get_current_user()
            if not user:
                return self._send_json({"error": "Sesion requerida."}, status=401)
            return self._send_json({"user": user})
        if parsed.path == "/api/auth/users":
            try:
                self._require_roles({"admin"})
                return self._send_json(self.auth.list_admin_users())
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)
        if parsed.path == "/api/flota/catalogos":
            try:
                self._require_roles({"admin", "supervisor", "recepcion"})
                sucursal_scope = self._get_flota_sucursal_scope()
                return self._send_json(self.repo.get_flota_catalogos(sucursal=sucursal_scope))
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)
        if parsed.path == "/api/flota/vehiculos":
            try:
                self._require_roles({"admin", "supervisor", "recepcion"})
                activo = query.get("activo", [None])[0]
                activo_bool = None if activo is None or activo == "" else _parse_bool(activo)
                sucursal = self._get_flota_sucursal_scope(query.get("sucursal", [None])[0])
                return self._send_json(self.repo.list_vehiculos(activo=activo_bool, sucursal=sucursal))
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)
        if parsed.path == "/api/flota/combustible":
            try:
                self._require_roles({"admin", "supervisor", "recepcion"})
                sucursal_scope = self._get_flota_sucursal_scope(query.get("sucursal", [None])[0])
                return self._send_json(
                    self.repo.list_cargas_combustible(
                        desde=_parse_date(query.get("desde", [None])[0]),
                        hasta=_parse_date(query.get("hasta", [None])[0]),
                        vehiculo_id=query.get("vehiculo_id", [None])[0],
                        sucursal=sucursal_scope,
                    )
                )
            except ValueError as exc:
                return self._send_json({"error": str(exc)}, status=400)
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)
        if parsed.path == "/api/flota/gastos":
            try:
                self._require_roles({"admin", "supervisor", "recepcion"})
                sucursal_scope = self._get_flota_sucursal_scope(query.get("sucursal", [None])[0])
                return self._send_json(
                    self.repo.list_gastos_flota(
                        desde=_parse_date(query.get("desde", [None])[0]),
                        hasta=_parse_date(query.get("hasta", [None])[0]),
                        vehiculo_id=query.get("vehiculo_id", [None])[0],
                        tipo_gasto_id=query.get("tipo_gasto_id", [None])[0],
                        sucursal=sucursal_scope,
                    )
                )
            except ValueError as exc:
                return self._send_json({"error": str(exc)}, status=400)
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)
        if parsed.path == "/api/flota/resumen-semanal":
            try:
                self._require_roles({"admin", "supervisor", "recepcion"})
                sucursal_scope = self._get_flota_sucursal_scope(query.get("sucursal", [None])[0])
                return self._send_json(
                    self.repo.get_flota_resumen_semanal(
                        semana=query.get("mes", [query.get("semana", [None])[0]])[0],
                        anho=query.get("anho", [None])[0],
                        vehiculo_id=query.get("vehiculo_id", [None])[0],
                        sucursal=sucursal_scope,
                    )
                )
            except ValueError as exc:
                return self._send_json({"error": str(exc)}, status=400)
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)
        if parsed.path == "/api/flota/resumen-mensual/pdf":
            try:
                self._require_roles({"admin", "supervisor", "recepcion"})
                sucursal_scope = self._get_flota_sucursal_scope(query.get("sucursal", [None])[0])
                pdf_bytes, filename = self.repo.build_flota_resumen_mensual_pdf(
                    mes=query.get("mes", [None])[0],
                    anho=query.get("anho", [None])[0],
                    vehiculo_id=query.get("vehiculo_id", [None])[0],
                    sucursal=sucursal_scope,
                )
                return self._send_pdf(pdf_bytes, filename)
            except ValueError as exc:
                return self._send_json({"error": str(exc)}, status=400)
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)

        if parsed.path == "/api/resumenes":
            try:
                self._require_roles({"admin", "supervisor"})
                lote_ids = query.get("lote_ids", [None])[0]
                return self._send_json(self.repo.get_resumenes_data(lote_ids=lote_ids))
            except ValueError as exc:
                return self._send_json({"error": str(exc)}, status=400)
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)

        if parsed.path == "/api/estadisticas":
            try:
                self._require_roles({"admin", "supervisor"})
                desde = _parse_date(query.get("desde", [None])[0])
                hasta = _parse_date(query.get("hasta", [None])[0])
                return self._send_json(self.repo.get_estadisticas_data(desde=desde, hasta=hasta))
            except ValueError as exc:
                return self._send_json({"error": str(exc)}, status=400)
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)

        if parsed.path == "/api/resumenes/pdf":
            try:
                self._require_roles({"admin", "supervisor"})
                lote_ids = query.get("lote_ids", [None])[0]
                pdf_bytes = self.repo.build_resumenes_pdf(lote_ids)
                return self._send_pdf(pdf_bytes, "Reporte_Lotes_Completados.pdf")
            except ValueError as exc:
                return self._send_json({"error": str(exc)}, status=400)
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)

        if parsed.path == "/api/compras-faena":
            try:
                self._require_roles({"admin", "supervisor"})
                lote_id = query.get("lote_id", [None])[0]
                return self._send_json(self.repo.get_compras_faena_data(lote_id=lote_id))
            except ValueError as exc:
                return self._send_json({"error": str(exc)}, status=400)
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)

        if parsed.path == "/api/distribuciones":
            try:
                self._require_roles({"admin", "supervisor"})
                lote_id = query.get("lote_id", [None])[0]
                return self._send_json(self.repo.get_distribuciones_data(lote_id=lote_id))
            except ValueError as exc:
                return self._send_json({"error": str(exc)}, status=400)
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)

        recepcion_slug, recepcion_action = _recepcion_slug_from_path(parsed.path)
        if recepcion_slug and recepcion_action == "":
            try:
                self._require_sucursal_access(recepcion_slug)
                fecha = _parse_date(query.get("fecha", [None])[0]) or date.today()
                payload = self.repo.get_recepcion(recepcion_slug, fecha=fecha)
                return self._send_json(payload)
            except ValueError:
                return self._send_json({"error": "Sucursal o fecha invalida. Use YYYY-MM-DD."}, status=400)
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)

        if recepcion_slug and recepcion_action == "pdf":
            try:
                user = self._require_sucursal_access(recepcion_slug)
                fecha = _parse_date(query.get("fecha", [None])[0]) or date.today()
                pdf_bytes, filename = self.repo.build_recepcion_pdf(
                    recepcion_slug,
                    fecha=fecha,
                    generated_by=user.get("nombre") or user.get("username"),
                    user_role=user.get("rol"),
                    allowed_scope=user.get("sucursal_permitida"),
                )
                return self._send_pdf(pdf_bytes, filename)
            except ValueError:
                return self._send_json({"error": "Sucursal o fecha invalida. Use YYYY-MM-DD."}, status=400)
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)

        if parsed.path != "/api/dashboard":
            return self._send_json({"error": "Ruta no encontrada"}, status=404)

        try:
            self._require_roles({"admin", "supervisor"})
            desde = _parse_date(query.get("desde", [None])[0])
            hasta = _parse_date(query.get("hasta", [None])[0])
            payload = self.repo.get_dashboard(desde=desde, hasta=hasta)
            return self._send_json(payload)
        except ValueError:
            return self._send_json({"error": "Formato de fecha invalido. Use YYYY-MM-DD."}, status=400)
        except PermissionDenied as exc:
            return self._send_json({"error": str(exc)}, status=403)
        except AuthError as exc:
            return self._send_json({"error": str(exc)}, status=401)
        except Exception as exc:
            return self._send_json({"error": str(exc)}, status=500)

    def do_POST(self):
        parsed = urlparse(self.path)
        query = parse_qs(parsed.query)
        if parsed.path == "/api/contratos":
            try:
                from web.backend.modules.contratos.repository import ContratosRepository

                self._require_module_access("contratos")
                payload = self._read_json()
                return self._send_json(
                    ContratosRepository(DATABASE_URL).save_contrato(payload),
                    status=201 if not payload.get("id") else 200,
                )
            except ValueError as exc:
                return self._send_json({"error": str(exc)}, status=400)
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)
        if parsed.path == "/api/acuerdos-comerciales/cobranzas":
            try:
                payload = self._read_json()
                from web.backend.modules.acuerdos_comerciales.repository import AcuerdosComercialesRepository
                from web.backend.modules.acuerdos_comerciales.schemas import ROLES_ACUERDOS

                user = self._require_roles(ROLES_ACUERDOS)
                return self._send_json(
                    AcuerdosComercialesRepository(DATABASE_URL).save_acuerdo_cobranza(
                        payload,
                        cambiado_por=user.get("username"),
                    ),
                    status=200,
                )
            except ValueError as exc:
                return self._send_json({"error": str(exc)}, status=400)
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)
        if parsed.path == "/api/acuerdos-comerciales/mapa/ubicaciones/valor":
            try:
                payload = self._read_json()
                from web.backend.modules.acuerdos_comerciales.repository import AcuerdosComercialesRepository
                from web.backend.modules.acuerdos_comerciales.schemas import ROLES_ACUERDOS

                user = self._require_roles(ROLES_ACUERDOS)
                modules = user.get("modulos_permitidos")
                if isinstance(modules, list) and "acuerdos-valores" not in modules:
                    raise PermissionDenied("No tienes permisos para cambiar valores de ubicaciones.")
                return self._send_json(
                    AcuerdosComercialesRepository(DATABASE_URL).save_mapa_ubicacion_valor(
                        payload,
                        cambiado_por=user.get("username"),
                    ),
                    status=200,
                )
            except ValueError as exc:
                return self._send_json({"error": str(exc)}, status=400)
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)
        if parsed.path.rstrip("/") == "/api/acuerdos-comerciales/eliminar":
            try:
                payload = self._read_json()
                from web.backend.modules.acuerdos_comerciales.routes import service as acuerdos_service
                from web.backend.modules.acuerdos_comerciales.schemas import ROLES_ACUERDOS

                user = self._require_roles(ROLES_ACUERDOS)
                return self._send_json(acuerdos_service.eliminar_acuerdo(payload, cambiado_por=user.get("username")), status=200)
            except ValueError as exc:
                return self._send_json({"error": str(exc)}, status=400)
            except PermissionDenied as exc:
                return self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                return self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                return self._send_json({"error": str(exc)}, status=500)
        if self._has_module_route("POST", parsed.path):
            try:
                payload = self._read_json()
            except ValueError as exc:
                return self._send_json({"error": str(exc)}, status=400)
            if self._dispatch_module_route("POST", parsed, query, payload):
                return
        try:
            payload = self._read_json()
            if parsed.path == "/api/auth/login":
                user, token, max_age = self.auth.login(
                    payload.get("username"),
                    payload.get("password"),
                    ip=self.client_address[0] if self.client_address else None,
                    user_agent=self.headers.get("User-Agent"),
                )
                return self._send_json(
                    {"ok": True, "user": user},
                    status=200,
                    extra_headers=[("Set-Cookie", self._build_session_cookie(token, max_age))],
                )
            if parsed.path == "/api/auth/logout":
                token = self._get_session_token()
                self.auth.logout(token)
                return self._send_json(
                    {"ok": True},
                    status=200,
                    extra_headers=[("Set-Cookie", self._build_logout_cookie())],
                )
            if parsed.path == "/api/auth/users":
                self._require_roles({"admin"})
                return self._send_json(
                    {
                        "user": self.auth.create_admin_user_with_scope(
                            payload.get("username"),
                            payload.get("nombre"),
                            payload.get("password"),
                            payload.get("rol"),
                            payload.get("sucursal_permitida"),
                            bool(payload.get("activo", True)),
                            payload.get("modulos_permitidos"),
                        )
                    },
                    status=201,
                )
            if parsed.path == "/api/flota/vehiculos":
                self._require_roles({"admin", "supervisor"})
                sucursal_scope = self._get_flota_sucursal_scope(payload.get("sucursal"))
                return self._send_json(
                    self.repo.save_vehiculo(payload, sucursal_scope=sucursal_scope),
                    status=201 if not payload.get("id") else 200,
                )
            if parsed.path == "/api/flota/proveedores":
                self._require_roles({"admin", "supervisor"})
                return self._send_json(
                    self.repo.save_proveedor_flota(payload),
                    status=201 if not payload.get("id") else 200,
                )
            if parsed.path == "/api/flota/combustible":
                user = self._require_roles({"admin", "supervisor", "recepcion"})
                sucursal_scope = self._get_flota_sucursal_scope()
                return self._send_json(
                    self.repo.save_carga_combustible(payload, cargado_por=user.get("username"), sucursal_scope=sucursal_scope),
                    status=201,
                )
            if parsed.path == "/api/flota/combustible/eliminar":
                user = self._require_roles({"admin", "supervisor", "recepcion"})
                sucursal_scope = self._get_flota_sucursal_scope()
                return self._send_json(
                    self.repo.delete_carga_combustible(payload, eliminado_por=user.get("username"), sucursal_scope=sucursal_scope)
                )
            if parsed.path == "/api/flota/combustible/import":
                user = self._require_roles({"admin", "supervisor"})
                sucursal_scope = self._get_flota_sucursal_scope()
                return self._send_json(
                    self.repo.import_cargas_combustible(payload, cargado_por=user.get("username"), sucursal_scope=sucursal_scope),
                    status=201,
                )
            if parsed.path == "/api/flota/combustible/import/preview":
                self._require_roles({"admin", "supervisor"})
                sucursal_scope = self._get_flota_sucursal_scope()
                return self._send_json(self.repo.preview_cargas_combustible_import(payload, sucursal_scope=sucursal_scope))
            if parsed.path == "/api/flota/gastos":
                user = self._require_roles({"admin", "supervisor", "recepcion"})
                sucursal_scope = self._get_flota_sucursal_scope()
                return self._send_json(
                    self.repo.save_gasto_flota(payload, cargado_por=user.get("username"), sucursal_scope=sucursal_scope, user_role=user.get("rol")),
                    status=201 if not payload.get("id") else 200,
                )
            if parsed.path == "/api/flota/gastos/eliminar":
                user = self._require_roles({"admin", "supervisor", "recepcion"})
                sucursal_scope = self._get_flota_sucursal_scope()
                return self._send_json(
                    self.repo.delete_gasto_flota(payload, eliminado_por=user.get("username"), sucursal_scope=sucursal_scope, user_role=user.get("rol"))
                )
            if parsed.path == "/api/compras-faena/lotes":
                self._require_roles({"admin", "supervisor"})
                return self._send_json(self.repo.save_lote(payload), status=201 if not payload.get("id") else 200)
            if parsed.path == "/api/compras-faena/lotes/eliminar":
                self._require_roles({"admin", "supervisor"})
                return self._send_json(self.repo.delete_lote_compra(payload))
            if parsed.path == "/api/compras-faena/faenas":
                self._require_roles({"admin", "supervisor"})
                return self._send_json(self.repo.add_faena(payload), status=201)
            if parsed.path == "/api/compras-faena/faena-total":
                self._require_roles({"admin", "supervisor"})
                return self._send_json(self.repo.set_faena_total(payload))
            if parsed.path == "/api/resumenes/cerrar":
                self._require_roles({"admin", "supervisor"})
                return self._send_json(self.repo.set_lotes_resumen_cerrado(payload))
            if parsed.path == "/api/distribuciones":
                self._require_roles({"admin", "supervisor"})
                return self._send_json(self.repo.save_distribucion(payload))
            recepcion_slug, recepcion_action = _recepcion_slug_from_path(parsed.path)
            if recepcion_slug and recepcion_action == "distribuciones":
                self._require_roles({"admin", "supervisor", "recepcion"})
                self._require_sucursal_access(recepcion_slug)
                return self._send_json(self.repo.update_recepcion_distribucion(recepcion_slug, payload))
            if recepcion_slug and recepcion_action == "menudencias":
                self._require_roles({"admin", "supervisor", "recepcion"})
                self._require_sucursal_access(recepcion_slug)
                return self._send_json(self.repo.add_menudencia(recepcion_slug, payload), status=201)
            return self._send_json({"error": "Ruta no encontrada"}, status=404)
        except ValueError as exc:
            return self._send_json({"error": str(exc)}, status=400)
        except PermissionDenied as exc:
            return self._send_json({"error": str(exc)}, status=403)
        except AuthError as exc:
            return self._send_json({"error": str(exc)}, status=401)
        except Exception as exc:
            return self._send_json({"error": str(exc)}, status=500)

    def do_PUT(self):
        parsed = urlparse(self.path)
        query = parse_qs(parsed.query)
        if self._has_module_route("PUT", parsed.path):
            try:
                payload = self._read_json()
            except ValueError as exc:
                return self._send_json({"error": str(exc)}, status=400)
            if self._dispatch_module_route("PUT", parsed, query, payload):
                return
        try:
            payload = self._read_json()
            if parsed.path == "/api/auth/users":
                self._require_roles({"admin"})
                return self._send_json(
                    {
                        "user": self.auth.update_admin_user_with_scope(
                            payload.get("id"),
                            payload.get("nombre"),
                            payload.get("rol"),
                            bool(payload.get("activo", True)),
                            payload.get("sucursal_permitida"),
                            payload.get("modulos_permitidos"),
                        )
                    }
                )
            if parsed.path == "/api/auth/users/password":
                self._require_roles({"admin"})
                self.auth.update_admin_password(payload.get("id"), payload.get("password"))
                return self._send_json({"ok": True})
            if parsed.path == "/api/flota/vehiculos":
                self._require_roles({"admin", "supervisor"})
                sucursal_scope = self._get_flota_sucursal_scope(payload.get("sucursal"))
                return self._send_json(self.repo.save_vehiculo(payload, sucursal_scope=sucursal_scope))
            if parsed.path == "/api/flota/proveedores":
                self._require_roles({"admin", "supervisor"})
                return self._send_json(self.repo.save_proveedor_flota(payload))
            recepcion_slug, recepcion_action = _recepcion_slug_from_path(parsed.path)
            if recepcion_slug and recepcion_action == "menudencias":
                self._require_roles({"admin", "supervisor", "recepcion"})
                self._require_sucursal_access(recepcion_slug)
                return self._send_json(self.repo.update_menudencia(recepcion_slug, payload))
            return self._send_json({"error": "Ruta no encontrada"}, status=404)
        except ValueError as exc:
            return self._send_json({"error": str(exc)}, status=400)
        except PermissionDenied as exc:
            return self._send_json({"error": str(exc)}, status=403)
        except AuthError as exc:
            return self._send_json({"error": str(exc)}, status=401)
        except Exception as exc:
            return self._send_json({"error": str(exc)}, status=500)

    def do_PATCH(self):
        self._send_json({"error": "Metodo no habilitado"}, status=405)

    def do_DELETE(self):
        parsed = urlparse(self.path)
        query = parse_qs(parsed.query)
        if self._dispatch_module_route("DELETE", parsed, query):
            return
        try:
            if parsed.path == "/api/acuerdos-comerciales":
                from web.backend.modules.acuerdos_comerciales.routes import service as acuerdos_service
                from web.backend.modules.acuerdos_comerciales.schemas import ROLES_ACUERDOS

                user = self._require_roles(ROLES_ACUERDOS)
                acuerdo_id = query.get("id", [None])[0] or query.get("acuerdo_id", [None])[0]
                if not acuerdo_id:
                    raise ValueError("El acuerdo_id es obligatorio.")
                return self._send_json(
                    acuerdos_service.descartar_negociacion(acuerdo_id, cambiado_por=user.get("username")),
                    status=200,
                )
            if parsed.path == "/api/distribuciones":
                self._require_roles({"admin", "supervisor"})
                query = parse_qs(parsed.query)
                distrib_id = query.get("id", [None])[0]
                if not distrib_id:
                    raise ValueError("Falta id de distribucion.")
                return self._send_json(self.repo.delete_distribucion(distrib_id))
            recepcion_slug, recepcion_action = _recepcion_slug_from_path(parsed.path)
            if recepcion_slug and recepcion_action == "menudencias":
                self._require_roles({"admin", "supervisor", "recepcion"})
                self._require_sucursal_access(recepcion_slug)
                query = parse_qs(parsed.query)
                men_id = query.get("id", [None])[0]
                if not men_id:
                    raise ValueError("Falta id de menudencia.")
                return self._send_json(self.repo.delete_menudencia(recepcion_slug, men_id))
            return self._send_json({"error": "Ruta no encontrada"}, status=404)
        except ValueError as exc:
            return self._send_json({"error": str(exc)}, status=400)
        except PermissionDenied as exc:
            return self._send_json({"error": str(exc)}, status=403)
        except AuthError as exc:
            return self._send_json({"error": str(exc)}, status=401)
        except Exception as exc:
            return self._send_json({"error": str(exc)}, status=500)

    def _send_empty(self, status=204, extra_headers: list[tuple[str, str]] | None = None):
        try:
            self.send_response(status)
            self._headers()
            for key, value in extra_headers or []:
                self.send_header(key, value)
            self.end_headers()
        except OSError as exc:
            if _is_client_disconnect(exc):
                return
            raise

    def _send_json(self, payload, status=200, extra_headers: list[tuple[str, str]] | None = None):
        body = json.dumps(payload, default=_json_default).encode("utf-8")
        try:
            self.send_response(status)
            self._headers()
            for key, value in extra_headers or []:
                self.send_header(key, value)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
            self.send_header("Pragma", "no-cache")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        except OSError as exc:
            if _is_client_disconnect(exc):
                return
            raise

    def _send_pdf(self, body: bytes, filename: str, status=200, extra_headers: list[tuple[str, str]] | None = None):
        try:
            self.send_response(status)
            self._headers()
            for key, value in extra_headers or []:
                self.send_header(key, value)
            self.send_header("Content-Type", "application/pdf")
            self.send_header("Content-Disposition", f'inline; filename="{filename}"')
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        except OSError as exc:
            if _is_client_disconnect(exc):
                return
            raise

    def _read_json(self):
        length = int(self.headers.get("Content-Length") or 0)
        if length <= 0:
            return {}
        raw = self.rfile.read(length).decode("utf-8")
        return json.loads(raw)

    def _has_module_route(self, method: str, path: str) -> bool:
        return any(route.matches(method, path) for route in self.module_routes)

    def _dispatch_module_route(self, method: str, parsed, query, payload=None) -> bool:
        for route in self.module_routes:
            if not route.matches(method, parsed.path):
                continue
            try:
                result = route.handler(RequestContext(self, parsed, query, payload))
                if result is not None:
                    self._send_json(result)
                return True
            except ValueError as exc:
                self._send_json({"error": str(exc)}, status=400)
            except PermissionDenied as exc:
                self._send_json({"error": str(exc)}, status=403)
            except AuthError as exc:
                self._send_json({"error": str(exc)}, status=401)
            except Exception as exc:
                logging.exception("Error no manejado en ruta de modulo %s %s", method, parsed.path)
                self._send_json({"error": str(exc)}, status=500)
            return True

    def _headers(self):
        origin = self.headers.get("Origin", "")
        allowed_origins = {
            "http://localhost:4200",
            "http://127.0.0.1:4200",
            "http://192.168.10.12:4200",
        }
        self.send_header(
            "Access-Control-Allow-Origin",
            origin if origin in allowed_origins else "http://127.0.0.1:4200",
        )
        self.send_header("Access-Control-Allow-Methods", "GET, POST, PUT, DELETE, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Access-Control-Allow-Credentials", "true")
        self.send_header("Vary", "Origin")

    def _get_session_token(self) -> str | None:
        raw_cookie = self.headers.get("Cookie", "")
        if not raw_cookie:
            return None
        cookie = SimpleCookie()
        cookie.load(raw_cookie)
        morsel = cookie.get("reces_session")
        return morsel.value if morsel else None

    def _get_current_user(self) -> dict[str, Any] | None:
        return self.auth.get_current_user(self._get_session_token())

    def _require_roles(self, allowed_roles: set[str]) -> dict[str, Any]:
        return self.auth.require_roles(self._get_current_user(), allowed_roles)

    def _require_module_access(self, module_key: str) -> dict[str, Any]:
        user = self._get_current_user()
        if not user:
            raise AuthError("Sesion requerida.")
        role = str(user.get("rol") or "")
        modules = user.get("modulos_permitidos")
        if isinstance(modules, list):
            if module_key not in modules:
                raise PermissionDenied("No tienes permisos para este modulo.")
            return user
        if role in {"admin", "supervisor"}:
            return user
        raise PermissionDenied("No tienes permisos para este modulo.")

    def _get_flota_sucursal_scope(self, requested_slug: str | None = None) -> str | None:
        user = self._get_current_user()
        if not user:
            raise AuthError("Sesion requerida.")
        role = str(user.get("rol") or "")
        requested = str(requested_slug or "").strip().lower() or None
        if role in {"admin", "supervisor"}:
            return requested
        if role == "recepcion":
            allowed_slug = str(user.get("sucursal_permitida") or "").strip().lower()
            if not allowed_slug:
                raise PermissionDenied("Tu usuario no tiene una sucursal permitida configurada.")
            if requested and requested != allowed_slug:
                raise PermissionDenied("No tienes permisos para esta sucursal.")
            return allowed_slug
        raise PermissionDenied("No tienes permisos para flota.")

    def _require_sucursal_access(self, slug: str) -> dict[str, Any]:
        user = self._get_current_user()
        if not user:
            raise AuthError("Sesion requerida.")
        role = str(user.get("rol") or "")
        if role in {"admin", "supervisor"}:
            return user
        allowed_slug = str(user.get("sucursal_permitida") or "").strip().lower()
        if role == "recepcion" and allowed_slug == (slug or "").strip().lower():
            return user
        raise PermissionDenied("No tienes permisos para esta sucursal.")

    def _build_session_cookie(self, token: str, max_age: int) -> str:
        return f"reces_session={token}; HttpOnly; Path=/; Max-Age={max_age}; SameSite=Lax"

    def _build_logout_cookie(self) -> str:
        return "reces_session=; HttpOnly; Path=/; Max-Age=0; SameSite=Lax"


def main():
    server = ThreadingHTTPServer((HOST, PORT), DashboardHandler)
    module_summary: dict[str, dict[str, Any]] = {}
    for route in DashboardHandler.module_routes:
        parts = route.path.strip("/").split("/")
        module_name = parts[1] if len(parts) > 1 else route.path
        summary = module_summary.setdefault(module_name, {"count": 0, "methods": set()})
        summary["count"] += 1
        summary["methods"].add(route.method)
    module_text = ", ".join(
        f"{name} ({'/'.join(sorted(summary['methods']))}: {summary['count']} rutas)"
        for name, summary in sorted(module_summary.items())
    )
    print(f"Dashboard API en http://{HOST}:{PORT}")
    print("Core: /api/health, /api/auth/login, /api/auth/logout, /api/auth/me, /api/auth/users")
    print("Legacy: /api/dashboard, /api/compras-faena, /api/distribuciones, /api/recepcion, /api/flota/*")
    print(f"Modulos auto-registrados: {module_text or 'ninguno'}")
    server.serve_forever()


if __name__ == "__main__":
    main()
